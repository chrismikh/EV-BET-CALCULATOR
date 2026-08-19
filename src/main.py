from __future__ import annotations

import sys, os, re
from datetime import datetime
from dataclasses import dataclass
from typing import List, Dict, Tuple, Optional

from PyQt6.QtCore import Qt, QThread, pyqtSignal, QObject, QTimer, QDate
from PyQt6.QtGui import QAction, QIcon, QColor
from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, QLabel,
    QTableWidget, QTableWidgetItem, QComboBox, QPushButton, QGroupBox, QLineEdit,
    QMessageBox, QDialog, QProgressBar, QStatusBar, QHeaderView,
    QTreeWidget, QTreeWidgetItem, QTabWidget, QFrame, QFileDialog,
    QFormLayout, QScrollArea, QInputDialog, QGridLayout, QDateEdit, QSizePolicy
)

# Ensure project root is in sys.path so we can import from src
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

try:
    import src.theme_manager as theme_manager
except ModuleNotFoundError:
    import theme_manager as theme_manager

try:
    from src.database import DatabaseManager
except ModuleNotFoundError:
    from database import DatabaseManager

# RowTuple now stores: (odds, live_wr, prematch_wr, live_bet_count, prematch_bet_count)
RowTuple = Tuple[float, Optional[float], Optional[float], Optional[int], Optional[int]]
MatchBetTuple = Tuple[str, str, str, str, str, float, str, str]


@dataclass
class SheetCacheEntry:
    rows: List[RowTuple]
    # index maps odds -> (live_wr, prem_wr, live_cnt, prem_cnt)
    index: Dict[float, Tuple[Optional[float], Optional[float], Optional[int], Optional[int]]]


def resource_path(name: str) -> str:
    if getattr(sys, "frozen", False):
        base = getattr(sys, "_MEIPASS", os.path.dirname(sys.executable))  # type: ignore[attr-defined]
    else:
        base = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(base, name)


def round_odds_key(v) -> Optional[float]:
    try:
        return round(float(v), 4)
    except Exception:
        return None


def calc_true_prob(odds: float, wr_raw: float, n: int, k: int = 50):
    """
    Compute the true probability using Bayesian shrinkage:
    p_final = (wr_raw * n + p_market * k) / (n + k)
    """
    if wr_raw is None or n is None:
        return None

    # Market implied probability
    p_market = 1.0 / odds

    # Bayesian shrinkage
    p_final = (wr_raw * n + p_market * k) / (n + k)
    return p_final


def calc_ev(odds: float, wr_raw: float, n: int):
    """
    EV = p_final * odds - 1
    """
    p = calc_true_prob(odds, wr_raw, n)
    if p is None:
        return None
    return p * odds - 1.0


def extract_team_from_bet(matchup: str, bet: str) -> Optional[str]:
    """
    Given a matchup string (e.g. 'Team A vs Team B') and a bet string
    (e.g. '2nd Map: +3.5 Team A'), determine which team the bet was placed on.
    Returns the matched team name or None if no team match is found.

    If the matchup has no 'vs' (e.g. 'Game 6', 'Main Event'), extracts the
    value after the last ': ' in the bet string (e.g. 'Winner: NRG' -> 'NRG',
    'Fastest lap: Verstappen' -> 'Verstappen').
    """
    if not bet:
        return None

    # Check if matchup contains 'vs'
    has_vs = matchup and re.search(r'\s+vs\s+', matchup, flags=re.IGNORECASE)

    if has_vs:
        # Split matchup on ' vs ' to get team names
        parts = re.split(r'\s+vs\s+', matchup, flags=re.IGNORECASE)
        if len(parts) != 2:
            return None

        team_a = parts[0].strip()
        team_b = parts[1].strip()

        if not team_a or not team_b:
            return None

        # Check which team name appears in the bet text
        a_in_bet = team_a.lower() in bet.lower()
        b_in_bet = team_b.lower() in bet.lower()

        if a_in_bet and b_in_bet:
            # Both match — prefer the longer name to avoid false positives
            return team_a if len(team_a) >= len(team_b) else team_b
        elif a_in_bet:
            return team_a
        elif b_in_bet:
            return team_b
        return None
    else:
        # No 'vs' in matchup — extract the value after the last ': ' in the bet
        if ': ' not in bet:
            return None
        after_colon = bet.rsplit(': ', 1)[1].strip()
        if not after_colon:
            return None
        # Strip leading handicap/number prefixes like '+3.5 ', '-1.5 '
        cleaned = re.sub(r'^[+-]?\d+(\.\d+)?\s+', '', after_colon)
        return cleaned if cleaned else None


def fmt_wr(wr: Optional[float]) -> str:
    return "N/A" if wr is None else f"{wr*100:.2f}%"


def fmt_ev(wr: Optional[float], odds: Optional[float], sample_size: Optional[int] = None):
    if wr is None or odds is None or sample_size is None:
        return "N/A", None

    ev = calc_ev(odds, wr, sample_size)
    if ev is None:
        return "N/A", None
    return f"{ev*100:.2f}%", ev


def fetch_matchbet_data_from_db(db: DatabaseManager, force_refresh_network: bool = False) -> List[MatchBetTuple]:
    """Fetch settled bets from the database.

    Returns data in the same MatchBetTuple format used elsewhere:
    (sport, tournament, matchup, bet, live_status, odds, result, provider)
    """
    try:
        return db.fetch_settled_bets(force_refresh=force_refresh_network)
    except Exception as e:
        print(f"Error fetching data from database: {e}")
        return []


def normalize_tournament_name(name: str) -> str:
    """
    Normalizes tournament names by removing years, seasons, and specific suffixes
    to group variations of the same event.
    """
    if not name:
        return ""

    # Keep Formula 1 grouped under a stable root label.
    if re.match(r'^Formula\s*1\b', name, re.IGNORECASE):
        return "Formula 1"

    # Map Esports World Cup / EWC variants to a single label
    if re.search(r'\b(Esports\s*World\s*Cup|EWC)\b', name, re.IGNORECASE):
        return "EWC"

    # Specific fix for StarLadder abbreviations
    name = name.replace("StarLadder SS", "StarLadder StarSeries")

    # --- VALORANT-specific normalizations ---
    # Fix common VCL/VLC typo
    name = re.sub(r'^VLC\b', 'VCL', name)
    # Unify VCT → Champions Tour
    name = re.sub(r'^VCT\b', 'Champions Tour', name)

    # VCT Regional Leagues (Americas, EMEA, Pacific, China)
    match = re.search(r'\bChampions Tour[\s\d:]*\b(Americas|EMEA|Pacific|China)\b', name, re.IGNORECASE)
    if match:
        region = match.group(1)
        if region.lower() == 'emea':
            return "VCT EMEA"
        return f"VCT {region.capitalize()}"

    # Game Changers (VALORANT) – detect before colon split loses the info
    if re.search(r'\bGame\s+Changers\b|:\s*GC\s', name, re.IGNORECASE):
        return "VCT GC"

    # Valorant Champions (world championship)
    if re.search(r'^Valorant\s+Champions\b', name, re.IGNORECASE):
        return "Valorant Champions"

    # VCT / Champions Tour Masters events
    if (re.search(r'^Champions Tour\b', name) and re.search(r'\bMasters\b', name, re.IGNORECASE)) \
            or re.match(r'^Masters\s+\w+', name, re.IGNORECASE):
        return "VCT Masters"

    # Champions Tour Ascension events
    if re.search(r'^Champions Tour\b', name) and re.search(r'\bAscension\b', name, re.IGNORECASE):
        return "Champions Tour Ascension"

    # VCL (Valorant Challengers League) – group all regions together
    if re.match(r'^VCL\b', name, re.IGNORECASE):
        return "VCL"

    # --- CS2-specific normalizations ---
    # IEM (Intel Extreme Masters) – group all city events together
    if re.match(r'^IEM\b', name):
        return "IEM"

    # PGL events – group all together
    if re.match(r'^PGL\b', name):
        return "PGL"

    # R6S BLAST Major events – group all host-city variants together
    if re.match(r'^BLAST\s+R6\s+Major\b', name, re.IGNORECASE):
        return "BLAST R6 Major"

    # Thunderpick / TP World Championship / TP WC – unify abbreviations
    if re.match(r'^(Thunderpick|TP)\s+(World\s+Championship|WC)\b', name, re.IGNORECASE):
        return "Thunderpick World Championship"

    # ESEA – group all variants (Advanced, divisions, etc.)
    if re.match(r'^ESEA\b', name):
        return "ESEA"

    # European Pro League – group all variants (Regular Season, Series, Division, etc.)
    if re.match(r'^European Pro League\b', name):
        return "European Pro League"

    # LCK – group all variants (Season, Challengers, etc.)
    if re.match(r'^LCK\b', name, re.IGNORECASE):
        return "LCK"

    # CS Asia Championships – preserve "Asia" before region stripping removes it
    if re.match(r'^CS\s+Asia\b', name, re.IGNORECASE):
        return "CS Asia Championships"

    # 1. Remove text after colons or double slashes (e.g., "BLAST: CQ" -> "BLAST", "Galaxy Battle // Phase 4" -> "Galaxy Battle")
    if '//' in name:
        name = name.split('//')[0]
    if ': ' in name:
        name = name.split(': ')[0]

    # 2. Remove years (1990-2029)
    # Matches 4 digits starting with 19 or 20 surrounded by word boundaries
    name = re.sub(r'\b(19|20)\d{2}\b', '', name)

    # 3. Remove common sequential patterns (Case insensitive)
    # "Season 19", "Series 5", "Vol. 2", "Part 1", "#12", "Stage 1", "Phase 2", "Split 1", "Group A"
    patterns = [
        r'\bSeason\s+\d+\b',
        r'\bSeries\s+\d+\b',
        r'\bVol\.?\s*\d+\b',
        r'\bPart\s+\d+\b',
        r'\bStage\s+\d+\b',
        r'\bPhase\s+\d+\b',
        r'\bSplit\s+\d+\b',
        r'\bGroup\s+[A-Za-z0-9]+\b',
        r'#\d+',
        r'\bOS\b',
        r'\b(Asia|Americas|Europe)\s+RMR(\s+[A-Z])?\b',
        r'\bRMR\b',
        r'\bS\d+\b',
        r'(?<!^)\b(Europe|EU|NA|SA|Asia|Americas|Oceania|CIS|European|South American|North American|North America|Pacific|APAC|EMEA|MENA|DACH)\b',
        r'\bLCQ\b',
        r'\b(Play-In|Global Finals|Contenders|CQ|Finals?|Groups?|Playoffs?)\b',
        r'\b\d+(?:st|nd|rd|th)?\s+Division\b',
        r'\bDivision\s+\d+\b',
        r'\bSeries\b',
        r'(?<!^)\b(Atlanta|Katowice|Bangkok|Raleigh|Lisbon)\b',
        r'\b(Spring|Summer|Fall|Winter)\b',
        r'\bWeek\b',
        r'\b(I|II|III|IV|V|VI|VII|VIII|IX|X)\b'
    ]
    for pattern in patterns:
        name = re.sub(pattern, '', name, flags=re.IGNORECASE)

    # 4. Remove standalone numbers (1-3 digits) that might be season/edition numbers
    # Only match numbers preceded by whitespace to preserve leading brand numbers (e.g. "500 Casino")
    name = re.sub(r'(?<=\s)\d{1,3}\b', '', name)

    # 5. Clean up extra whitespace and trailing hyphens
    # Replace multiple spaces with single space and trim ends
    name = re.sub(r'\s+', ' ', name).strip(' -')

    return name


def tournament_filter_labels(name: str) -> List[str]:
    """Return all tournament labels a row should match in the tournament filter."""
    base = normalize_tournament_name(name)
    labels: List[str] = [base] if base else []

    # Formula 1 supports both broad and race-specific filtering.
    if re.match(r'^Formula\s*1\b', name, re.IGNORECASE):
        race = re.sub(r'\b(19|20)\d{2}\b', '', name)
        race = re.sub(r'\s+', ' ', race).strip(' -')
        if ': ' in race:
            _, after = race.split(': ', 1)
            race = f"Formula 1: {after.strip()}"
        else:
            race = "Formula 1"

        if "Formula 1" not in labels:
            labels.insert(0, "Formula 1")
        if race and race not in labels:
            labels.append(race)

    return labels


def process_bets_to_cache(bets: List[MatchBetTuple]) -> Dict[str, SheetCacheEntry]:
    agg = {}
    for sport, _, _, _, live_status, odds, result, _provider in bets:
        if sport not in agg: agg[sport] = {}
        k = round_odds_key(odds)
        if k is None: continue
        if k not in agg[sport]: agg[sport][k] = {'lw':0,'lt':0,'pw':0,'pt':0}
        
        is_live = "LIVE" in live_status.upper() and "NOT" not in live_status.upper()
        is_win = result.lower() == "win"
        
        if is_live:
            agg[sport][k]['lt'] += 1
            if is_win: agg[sport][k]['lw'] += 1
        else:
            agg[sport][k]['pt'] += 1
            if is_win: agg[sport][k]['pw'] += 1
            
    cache = {}
    for sport, odds_map in agg.items():
        rows = []
        index = {}
        for k in sorted(odds_map.keys()):
            stats = odds_map[k]
            lt, lw = stats['lt'], stats['lw']
            pt, pw = stats['pt'], stats['pw']
            live_wr = (lw / lt) if lt > 0 else None
            prem_wr = (pw / pt) if pt > 0 else None
            tup = (k, live_wr, prem_wr, lt if lt > 0 else None, pt if pt > 0 else None)
            rows.append(tup)
            index[k] = (live_wr, prem_wr, lt if lt > 0 else None, pt if pt > 0 else None)
        cache[sport] = SheetCacheEntry(rows, index)
    return cache


class PreloadWorker(QObject):
    progress = pyqtSignal(str)
    status = pyqtSignal(str)
    finished = pyqtSignal(bool, str)
    def __init__(self, db: DatabaseManager):
        super().__init__()
        self.db = db
        self.cache: Dict[str, SheetCacheEntry] = {}
        self.matchbet_data: List[MatchBetTuple] = []

    def run(self):
        try:
            self.progress.emit("Loading data from database...")
            self.status.emit("Querying settled bets...")
            
            self.matchbet_data = fetch_matchbet_data_from_db(self.db)
            self.status.emit(f"Loaded {len(self.matchbet_data)} bets. Processing...")
            
            self.cache = process_bets_to_cache(self.matchbet_data)
            
            self.progress.emit("Finalizing...")
            self.finished.emit(True, "")
        except Exception as e:
            self.finished.emit(False, str(e))


class RefreshWorker(QObject):
    finished = pyqtSignal(bool, str, object)
    def __init__(self, db: DatabaseManager, force_refresh_network: bool = False):
        super().__init__()
        self.db = db
        self.force_refresh_network = force_refresh_network

    def run(self):
        try:
            data = fetch_matchbet_data_from_db(self.db, self.force_refresh_network)
            cache = process_bets_to_cache(data)
            self.finished.emit(True, "Refreshed all data", (cache, data))
        except Exception as e:
            self.finished.emit(False, str(e), None)


class PreloadDialog(QDialog):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Loading Data...")
        self.setFixedSize(420, 240)
        lay = QVBoxLayout(self)
        self.lbl_title = QLabel("Loading Betting Data")
        self.lbl_title.setAlignment(Qt.AlignmentFlag.AlignHCenter)
        self.lbl_title.setStyleSheet("font-size:18px;font-weight:bold;")
        lay.addWidget(self.lbl_title)
        self.lbl_progress = QLabel("Initializing...")
        lay.addWidget(self.lbl_progress)
        self.bar = QProgressBar(); self.bar.setRange(0,0); lay.addWidget(self.bar)
        self.lbl_status = QLabel("")
        self.lbl_status.setWordWrap(True)
        lay.addWidget(self.lbl_status)
    def update_progress(self, t: str): self.lbl_progress.setText(t)
    def update_status(self, t: str): self.lbl_status.setText(t)


class SortableTableWidgetItem(QTableWidgetItem):
    def __init__(self, text, sort_value):
        super().__init__(text)
        self.sort_value = sort_value

    def __lt__(self, other):
        v1 = self.sort_value
        v2 = other.sort_value
        if v1 is None: v1 = -float('inf')
        if v2 is None: v2 = -float('inf')
        return v1 < v2


class MigrationWorker(QObject):
    """Background worker to import an .xlsx file into the SQLite database."""
    progress = pyqtSignal(int, int)  # current, total
    finished = pyqtSignal(bool, str, int, int, int)  # ok, msg, total, settled, pending

    def __init__(self, file_path: str, db: DatabaseManager):
        super().__init__()
        self.file_path = file_path
        self.db = db

    def run(self):
        try:
            import openpyxl
        except ImportError:
            self.finished.emit(False, "openpyxl is not installed. Run: pip install openpyxl", 0, 0, 0)
            return
        try:
            wb = openpyxl.load_workbook(self.file_path, read_only=True, data_only=True)
            if "MATCHBET" not in wb.sheetnames:
                self.finished.emit(False, "Sheet 'MATCHBET' not found in the workbook.", 0, 0, 0)
                wb.close()
                return
            ws = wb["MATCHBET"]
            rows = list(ws.iter_rows(min_row=2, values_only=True))
            total = len(rows)
            db = self.db
            db.create_tables()
            imported = 0
            settled = 0
            pending = 0
            for i, row in enumerate(rows):
                self.progress.emit(i + 1, total)
                if len(row) < 6:
                    continue
                sport = str(row[0] or "").strip()
                if not sport:
                    continue
                tournament = str(row[1] or "").strip()
                matchup = str(row[2] or "").strip()
                bet = str(row[3] or "").strip()
                live_status = str(row[4] or "").strip() if len(row) > 4 else "NOT LIVE"
                odds_raw = row[5] if len(row) > 5 else None
                bet_amount_raw = row[6] if len(row) > 6 else None
                result_raw = str(row[7] or "").strip() if len(row) > 7 else ""
                profit_raw = row[8] if len(row) > 8 else None
                provider_raw = str(row[9] or "").strip() if len(row) > 9 else ""
                provider = provider_raw or "BetBy"

                try:
                    odds = float(str(odds_raw).replace(",", ".")) if odds_raw is not None else None
                except (ValueError, TypeError):
                    odds = None
                if odds is None or odds <= 0:
                    continue

                try:
                    bet_amount = float(str(bet_amount_raw).replace(",", ".")) if bet_amount_raw not in (None, "") else None
                except (ValueError, TypeError):
                    bet_amount = None

                result = result_raw if result_raw else None
                profit = None
                if result:
                    try:
                        profit = float(str(profit_raw).replace(",", ".")) if profit_raw not in (None, "") else None
                    except (ValueError, TypeError):
                        profit = None
                    settled += 1
                else:
                    pending += 1

                db.insert_bet(
                    sport=sport,
                    tournament=tournament,
                    matchup=matchup,
                    bet=bet,
                    live_status=live_status if live_status else "NOT LIVE",
                    provider=provider,
                    odds=odds,
                    bet_amount=bet_amount,
                    result=result,
                    profit=profit,
                )
                imported += 1
            wb.close()
            self.finished.emit(True, "", imported, settled, pending)
        except Exception as e:
            self.finished.emit(False, str(e), 0, 0, 0)


class SettingsDialog(QDialog):
    """Settings dialog with Appearance and Data Migration tabs."""

    def __init__(self, parent: "MainWindow"):
        super().__init__(parent)
        self.main_window = parent
        self.setWindowTitle("Settings")
        self.setFixedSize(650, 560)
        self.setModal(True)
        self._selected_file: Optional[str] = None
        self._migration_thread: Optional[QThread] = None
        self._build_ui()

    def _build_ui(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(12, 12, 12, 12)
        root.setSpacing(10)

        # Tab widget
        self.tabs = QTabWidget()
        root.addWidget(self.tabs, 1)

        # --- Tab 1: Appearance ---
        appearance_tab = QWidget()
        a_layout = QVBoxLayout(appearance_tab)
        a_layout.setContentsMargins(16, 16, 16, 16)
        a_layout.setSpacing(12)

        a_title = QLabel("Theme")
        a_title.setStyleSheet("font-size: 15px; font-weight: bold;")
        a_layout.addWidget(a_title)

        self.lbl_current_theme = QLabel(self._theme_status_text())
        self.lbl_current_theme.setStyleSheet("font-size: 13px;")
        a_layout.addWidget(self.lbl_current_theme)

        self.btn_theme = QPushButton()
        self.btn_theme.setCheckable(True)
        self.btn_theme.setChecked(self.main_window.dark_mode)
        self._sync_theme_button()
        self.btn_theme.toggled.connect(self._on_theme_toggled)
        a_layout.addWidget(self.btn_theme)

        a_layout.addStretch(1)
        self.tabs.addTab(appearance_tab, "Appearance")

        # --- Tab 2: Data Migration ---
        migration_tab = QWidget()
        m_layout = QVBoxLayout(migration_tab)
        m_layout.setContentsMargins(16, 16, 16, 16)
        m_layout.setSpacing(10)

        m_title = QLabel("Import Data from Excel (.xlsx)")
        m_title.setStyleSheet("font-size: 15px; font-weight: bold;")
        m_layout.addWidget(m_title)

        # Format instructions
        fmt_info = QLabel(
            "\U0001f4cb <b>Google Sheet / Excel Format Requirements:</b><br><br>"
            "Your file must have a sheet named <b>MATCHBET</b> with columns (in order):<br>"
            "&nbsp;&nbsp;A: Sport &nbsp; B: Tournament &nbsp; C: Matchup &nbsp; D: Bet<br>"
            "&nbsp;&nbsp;E: Live Status (LIVE or NOT LIVE) &nbsp; F: Odds<br>"
            "&nbsp;&nbsp;G: Bet Amount &nbsp; H: Result (Win/Lose) &nbsp; I: Profit<br>"
            "&nbsp;&nbsp;J: Provider (optional, defaults to BetBy)<br><br>"
            "\u26a0\ufe0f First row = headers (skipped). Empty Sport rows skipped.<br>"
            "\u26a0\ufe0f Bets without Result are imported as pending."
        )
        fmt_info.setWordWrap(True)
        fmt_info.setStyleSheet("font-size: 12px;")
        m_layout.addWidget(fmt_info)

        # Drag-and-drop area
        self.drop_frame = QFrame()
        self.drop_frame.setFrameShape(QFrame.Shape.StyledPanel)
        self.drop_frame.setMinimumHeight(80)
        self.drop_frame.setStyleSheet(
            "QFrame { border: 2px dashed #6c7086; border-radius: 10px; }"
            "QFrame:hover { border-color: #89b4fa; }"
        )
        self.drop_frame.setAcceptDrops(True)
        self.drop_frame.dragEnterEvent = self._drag_enter
        self.drop_frame.dropEvent = self._drop_event
        drop_layout = QVBoxLayout(self.drop_frame)
        drop_layout.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.lbl_drop = QLabel(
            "Drag and drop your .xlsx file here\nor click Browse below"
        )
        self.lbl_drop.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.lbl_drop.setStyleSheet("border: none; color: #6c7086; font-size: 13px;")
        drop_layout.addWidget(self.lbl_drop)
        m_layout.addWidget(self.drop_frame)

        # Browse button
        self.btn_browse = QPushButton("Browse Files")
        self.btn_browse.clicked.connect(self._browse_file)
        m_layout.addWidget(self.btn_browse)

        # Selected file label
        self.lbl_selected_file = QLabel("No file selected")
        self.lbl_selected_file.setStyleSheet("font-size: 12px;")
        self.lbl_selected_file.setWordWrap(True)
        m_layout.addWidget(self.lbl_selected_file)

        # Start Migration button
        self.btn_migrate = QPushButton("Start Migration")
        self.btn_migrate.setEnabled(False)
        self.btn_migrate.clicked.connect(self._start_migration)
        m_layout.addWidget(self.btn_migrate)

        # Progress bar (hidden initially)
        self.migration_progress = QProgressBar()
        self.migration_progress.setRange(0, 100)
        self.migration_progress.hide()
        m_layout.addWidget(self.migration_progress)

        # Status label
        self.lbl_migration_status = QLabel("")
        self.lbl_migration_status.setWordWrap(True)
        m_layout.addWidget(self.lbl_migration_status)

        m_layout.addStretch(1)
        self.tabs.addTab(migration_tab, "Data Migration")

        # --- Tab 3: Database Reset ---
        reset_tab = QWidget()
        r_layout = QVBoxLayout(reset_tab)
        r_layout.setContentsMargins(16, 16, 16, 16)
        r_layout.setSpacing(12)

        r_title = QLabel("Reset Database")
        r_title.setStyleSheet("font-size: 15px; font-weight: bold;")
        r_layout.addWidget(r_title)

        r_desc = QLabel(
            "This will permanently delete <b>all bets</b> from the database "
            "(both settled and pending). This action cannot be undone."
        )
        r_desc.setWordWrap(True)
        r_desc.setStyleSheet("font-size: 13px;")
        r_layout.addWidget(r_desc)

        self.btn_reset_db = QPushButton("Delete All Data")
        self.btn_reset_db.setStyleSheet(
            "QPushButton { background-color: #e74c3c; color: white; font-weight: bold; }"
            "QPushButton:hover { background-color: #c0392b; }"
        )
        self.btn_reset_db.clicked.connect(self._reset_database)
        r_layout.addWidget(self.btn_reset_db)

        self.lbl_reset_status = QLabel("")
        self.lbl_reset_status.setWordWrap(True)
        r_layout.addWidget(self.lbl_reset_status)

        r_layout.addStretch(1)
        self.tabs.addTab(reset_tab, "Database")

        # --- Close button ---
        btn_close = QPushButton("Close")
        btn_close.clicked.connect(self.accept)
        close_row = QHBoxLayout()
        close_row.addStretch(1)
        close_row.addWidget(btn_close)
        root.addLayout(close_row)

    # -- helpers --
    def _theme_status_text(self) -> str:
        mode = "Dark Mode" if self.main_window.dark_mode else "Light Mode"
        return f"Current Theme: {mode}"

    def _sync_theme_button(self):
        dark = self.main_window.dark_mode
        self.btn_theme.setText(" Light Mode" if dark else " Dark Mode")
        self.btn_theme.setIcon(
            QIcon(resource_path("icons/moon.svg" if dark else "icons/sun.svg"))
        )
        self.btn_theme.setToolTip("Toggle Dark / Light theme")

    def _on_theme_toggled(self, checked: bool):
        self.main_window.on_theme_toggled(checked)
        self.lbl_current_theme.setText(self._theme_status_text())
        self._sync_theme_button()

    # -- file handling --
    def _browse_file(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "Select Excel File", "", "Excel Files (*.xlsx)"
        )
        if path:
            self._set_selected_file(path)

    def _set_selected_file(self, path: str):
        self._selected_file = path
        name = os.path.basename(path)
        self.lbl_selected_file.setText(f"{name}\n{path}")
        self.btn_migrate.setEnabled(True)

    def _drag_enter(self, event):
        if event.mimeData().hasUrls():
            event.acceptProposedAction()

    def _drop_event(self, event):
        for url in event.mimeData().urls():
            path = url.toLocalFile()
            if path.lower().endswith(".xlsx"):
                self._set_selected_file(path)
                break

    # -- migration --
    def _start_migration(self):
        if not self._selected_file or not os.path.isfile(self._selected_file):
            QMessageBox.warning(self, "File Error", "Selected file does not exist.")
            return
        self.btn_migrate.setEnabled(False)
        self.btn_browse.setEnabled(False)
        self.migration_progress.setValue(0)
        self.migration_progress.show()
        self.lbl_migration_status.setText("Migrating...")

        self._migration_thread = QThread()
        self._migration_worker = MigrationWorker(self._selected_file, self.main_window.db)
        self._migration_worker.moveToThread(self._migration_thread)
        self._migration_thread.started.connect(self._migration_worker.run)
        self._migration_worker.progress.connect(self._on_migration_progress)
        self._migration_worker.finished.connect(self._on_migration_finished)
        self._migration_thread.start()

    def _on_migration_progress(self, current: int, total: int):
        if total > 0:
            self.migration_progress.setValue(int(current / total * 100))
        self.lbl_migration_status.setText(f"Processing row {current} / {total}...")

    def _on_migration_finished(self, ok: bool, msg: str, total: int, settled: int, pending: int):
        if self._migration_thread:
            self._migration_thread.quit()
            self._migration_thread.wait()
            self._migration_worker.deleteLater()
            self._migration_thread = None
        self.migration_progress.setValue(100 if ok else 0)
        self.btn_browse.setEnabled(True)
        if ok:
            self.lbl_migration_status.setText(
                f"Successfully migrated {total} bets ({settled} settled, {pending} pending)."
            )
            self.btn_migrate.setEnabled(False)
            # Trigger a data refresh on the main window
            self.main_window.refresh_data(force=True, force_network=True)
        else:
            self.lbl_migration_status.setText(f"Migration failed: {msg}")
            self.btn_migrate.setEnabled(True)

    # -- database reset --
    def _reset_database(self):
        reply = QMessageBox.warning(
            self,
            "Confirm Database Reset",
            "Are you sure you want to delete ALL bets from the database?\n\n"
            "This action cannot be undone.",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No,
        )
        if reply != QMessageBox.StandardButton.Yes:
            return
        try:
            db = self.main_window.db
            count = db.delete_all_bets()
            self.lbl_reset_status.setText(
                f"Deleted {count} bet(s). Database is now empty."
            )
            self.main_window.refresh_data(force=True)
        except Exception as e:
            self.lbl_reset_status.setText(f"Error: {e}")


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.db = DatabaseManager()
        self.db.create_tables()
        self.setWindowTitle("EV Bet Calculator")
        self.resize(1200, 800)
        self.setMinimumSize(1150, 600)
        self.data_cache: Dict[str, SheetCacheEntry] = {}
        self._previous_data_cache: Dict[str, SheetCacheEntry] = {}
        self.matchbet_data: List[MatchBetTuple] = []
        self._previous_matchbet_data: List[MatchBetTuple] = []
        self.current_rows: List[RowTuple] = []
        self.odds_index: Dict[float, Tuple[Optional[float], Optional[float], Optional[int], Optional[int]]] = {}
        self.dark_mode = True
        self.current_view = "table"  # "table", "statistics", "add_bet", "pending_bets"
        self.editing_bet_id: Optional[str] = None  # None = add mode, str = edit/settle mode
        self._statistics_panel_built = False
        self._statistics_refresh_pending = False
        self._build_ui()
        self._connect()
        theme_manager.apply_theme(QApplication.instance(), dark=self.dark_mode)

    # UI
    def _build_ui(self):
        # Central widget and root layout
        central = QWidget()
        self.setCentralWidget(central)
        root_layout = QHBoxLayout(central)

        # --- Sidebar (Controls) ---
        sidebar = QWidget(); sidebar.setObjectName("Sidebar"); sidebar.setFixedWidth(280)
        sidebar_layout = QVBoxLayout(sidebar)
        sidebar_layout.setContentsMargins(10, 10, 10, 10)
        sidebar_layout.setSpacing(10)
        root_layout.addWidget(sidebar)

        # Sport and Bet Type selectors
        controls_layout = QHBoxLayout()
        controls_layout.addWidget(QLabel("Sport:"))
        self.sport_combo = QComboBox(); controls_layout.addWidget(self.sport_combo)
        controls_layout.addWidget(QLabel("Bet Type:"))
        self.bettype_combo = QComboBox(); self.bettype_combo.addItems(["Live", "Not Live"]); controls_layout.addWidget(self.bettype_combo)
        sidebar_layout.addLayout(controls_layout)

        # Team selector
        team_layout = QHBoxLayout()
        team_layout.addWidget(QLabel("Team:"))
        self.team_combo = QComboBox()
        self.team_combo.addItem("All Teams")
        self.team_combo.setSizePolicy(self.team_combo.sizePolicy())
        team_layout.addWidget(self.team_combo, 1)
        sidebar_layout.addLayout(team_layout)

        # Tournament selector
        tournament_layout = QHBoxLayout()
        tournament_layout.addWidget(QLabel("Tournament:"))
        self.tournament_combo = QComboBox()
        self.tournament_combo.addItem("All Tournaments")
        self.tournament_combo.setSizePolicy(self.tournament_combo.sizePolicy())
        tournament_layout.addWidget(self.tournament_combo, 1)
        sidebar_layout.addLayout(tournament_layout)

        # Provider selector
        provider_layout = QHBoxLayout()
        provider_layout.addWidget(QLabel("Provider:"))
        self.provider_combo = QComboBox()
        self.provider_combo.addItem("All Providers")
        self.provider_combo.setSizePolicy(self.provider_combo.sizePolicy())
        provider_layout.addWidget(self.provider_combo, 1)
        sidebar_layout.addLayout(provider_layout)

        # Comparison GroupBox
        self.compare_group = QGroupBox("Comparison")
        cg = QVBoxLayout(self.compare_group)
        odds_a_layout = QHBoxLayout(); odds_a_layout.addWidget(QLabel("Odds A:")); self.entry_odds_a = QLineEdit(); odds_a_layout.addWidget(self.entry_odds_a); cg.addLayout(odds_a_layout)
        odds_b_layout = QHBoxLayout(); odds_b_layout.addWidget(QLabel("Odds B:")); self.entry_odds_b = QLineEdit(); odds_b_layout.addWidget(self.entry_odds_b); cg.addLayout(odds_b_layout)
        self.btn_compare = QPushButton("Compare"); cg.addWidget(self.btn_compare)
        sidebar_layout.addWidget(self.compare_group)
        
        # Statistics Button
        self.btn_statistics = QPushButton("Statistics")
        self.btn_statistics.setCheckable(True)
        sidebar_layout.addWidget(self.btn_statistics)
        
        # Data Table Button
        self.btn_data_table = QPushButton("Data Table")
        self.btn_data_table.setCheckable(True)
        self.btn_data_table.setChecked(True)  # default view
        sidebar_layout.addWidget(self.btn_data_table)

        # Add New Bet Button
        self.btn_add_bet = QPushButton("Add New Bet")
        self.btn_add_bet.setCheckable(True)
        sidebar_layout.addWidget(self.btn_add_bet)

        # Pending Bets Button
        self.btn_pending_bets = QPushButton("Pending Bets")
        self.btn_pending_bets.setCheckable(True)
        sidebar_layout.addWidget(self.btn_pending_bets)

        self._nav_buttons = [self.btn_statistics, self.btn_data_table, self.btn_add_bet, self.btn_pending_bets]
        
        sidebar_layout.addStretch(1)

        # Bottom buttons
        self.btn_refresh = QPushButton(" Force Refresh"); self.btn_refresh.setIcon(QIcon(resource_path("icons/refresh-cw.svg"))); sidebar_layout.addWidget(self.btn_refresh)

        # --- Main Content (Results) ---
        main_content = QWidget(); main_layout = QVBoxLayout(main_content); main_layout.setContentsMargins(10, 10, 10, 0); root_layout.addWidget(main_content, 1)

        # Comparison Result Card
        compare_card = QGroupBox(); compare_card.setObjectName("CompareCard"); card_layout = QVBoxLayout(compare_card)
        self.compare_result = QLabel("Enter two odds above and click Compare."); self.compare_result.setWordWrap(True); self.compare_result.setAlignment(Qt.AlignmentFlag.AlignCenter); self.compare_result.setMinimumHeight(80); card_layout.addWidget(self.compare_result); main_layout.addWidget(compare_card)

        # Data Table
        self.table = QTableWidget(0, 4)
        self.table.setHorizontalHeaderLabels(["Odds", "Live WR %", "Prematch WR %", "EV %"])
        
        self.table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.table.verticalHeader().setVisible(False)
        self.table.setSortingEnabled(True)

        # Configure Header
        header = self.table.horizontalHeader()
        header.setSortIndicatorShown(True)
        header.setSectionsMovable(False)
        header.setDefaultAlignment(Qt.AlignmentFlag.AlignCenter)
        
        # Stretch all columns to fill available space evenly
        header.setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        header.setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        header.setSectionResizeMode(2, QHeaderView.ResizeMode.Stretch)
        header.setSectionResizeMode(3, QHeaderView.ResizeMode.Stretch)
        
        main_layout.addWidget(self.table, 1)
        
        # Statistics Panel (initially hidden)
        self.statistics_panel = QWidget()
        statistics_layout = QVBoxLayout(self.statistics_panel)
        statistics_layout.setContentsMargins(10, 10, 10, 10)
        # Empty panel for now
        placeholder_label = QLabel("Statistics panel - coming soon")
        placeholder_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        placeholder_label.setStyleSheet("font-size: 16px; color: gray;")
        statistics_layout.addWidget(placeholder_label)
        main_layout.addWidget(self.statistics_panel, 1)
        self.statistics_panel.hide()  # Start hidden

        # --- Add Bet Panel (initially hidden) ---
        self._build_add_bet_panel(main_layout)

        # --- Pending Bets Panel (initially hidden) ---
        self._build_pending_bets_panel(main_layout)

        # --- Change Notification Panel (initially hidden) ---
        self.changes_panel = QWidget()
        changes_layout = QVBoxLayout(self.changes_panel)
        changes_layout.setContentsMargins(10, 10, 10, 10)
        
        # Header with title and dismiss button
        changes_header = QHBoxLayout()
        self.changes_title = QLabel("New Data Added")
        self.changes_title.setStyleSheet("font-size: 14px; font-weight: bold;")
        changes_header.addWidget(self.changes_title)
        changes_header.addStretch(1)
        self.btn_dismiss_changes = QPushButton("Dismiss")
        self.btn_dismiss_changes.setFixedWidth(80)
        changes_header.addWidget(self.btn_dismiss_changes)
        changes_layout.addLayout(changes_header)
        
        # Text area for changes
        self.changes_text = QLabel("No new data")
        self.changes_text.setWordWrap(True)
        self.changes_text.setAlignment(Qt.AlignmentFlag.AlignTop | Qt.AlignmentFlag.AlignLeft)
        self.changes_text.setTextInteractionFlags(Qt.TextInteractionFlag.TextSelectableByMouse)
        changes_layout.addWidget(self.changes_text)
        
        main_layout.addWidget(self.changes_panel, 0)
        self.changes_panel.hide()  # Start hidden

        # --- Status Bar & Menu ---
        self.status_bar = QStatusBar(); self.setStatusBar(self.status_bar); self.set_status("Ready")
        act_settings = QAction("Settings", self); act_settings.triggered.connect(self.open_settings_dialog); self.menuBar().addAction(act_settings)
        act_exit = QAction("Exit", self); act_exit.triggered.connect(self.close); self.menuBar().addAction(act_exit)

    def _connect(self):
        self.btn_refresh.clicked.connect(lambda checked=False: self.refresh_data(force=True, force_network=True))
        self.btn_compare.clicked.connect(self.compare_by_odds)
        self.sport_combo.currentTextChanged.connect(self.on_sport_change)
        self.bettype_combo.currentTextChanged.connect(self.on_bet_type_change)
        self.tournament_combo.currentTextChanged.connect(self.on_tournament_change)
        self.provider_combo.currentTextChanged.connect(self.on_provider_change)
        self.entry_odds_a.textChanged.connect(self.recompute_comparison_inline)
        self.entry_odds_b.textChanged.connect(self.recompute_comparison_inline)
        self.team_combo.currentTextChanged.connect(self.on_team_change)
        self.btn_statistics.clicked.connect(lambda checked=False: self.show_statistics_panel())
        self.btn_data_table.clicked.connect(self.show_data_table_panel)
        self.btn_add_bet.clicked.connect(self.show_add_bet_panel)
        self.btn_pending_bets.clicked.connect(self.show_pending_bets_panel)
        self.btn_dismiss_changes.clicked.connect(self.dismiss_changes_panel)

    # Helpers
    def set_status(self, text: str): self.status_bar.showMessage(text)
    def set_controls_enabled(self, enabled: bool):
        for w in [self.sport_combo,self.bettype_combo,self.team_combo,self.tournament_combo,self.provider_combo,self.btn_refresh,self.entry_odds_a,self.entry_odds_b,self.btn_compare]:
            w.setEnabled(enabled)

    def get_sorted_sports(self) -> List[str]:
        def count_bets(sport):
            entry = self.data_cache[sport]
            total = 0
            for r in entry.rows:
                # r[3] is live_cnt, r[4] is prem_cnt
                total += (r[3] or 0) + (r[4] or 0)
            return total
        return sorted(self.data_cache.keys(), key=count_bets, reverse=True)
    # Data
    def fill_table(self, rows: List[RowTuple], bet_type: str):
        self.table.setSortingEnabled(False)
        self.table.setRowCount(0)
        # Fetch theme colors from application properties (with fallbacks)
        app = QApplication.instance()
        from PyQt6.QtGui import QColor as _QColor  # local alias to avoid confusion
        pos = app.property("positiveColor") if app else None
        neg = app.property("negativeColor") if app else None
        neu = app.property("neutralColor") if app else None
        if not isinstance(pos, _QColor): pos = _QColor('green')
        if not isinstance(neg, _QColor): neg = _QColor('red')
        if not isinstance(neu, _QColor): neu = _QColor('gray')
        for odds, live_wr, prem_wr, _live_cnt, _prem_cnt in rows:
            wr = live_wr if bet_type=="Live" else prem_wr
            cnt = _live_cnt if bet_type=="Live" else _prem_cnt
            ev_str, ev_val = fmt_ev(wr, odds, cnt)
            r = self.table.rowCount(); self.table.insertRow(r)
            
            # Create sortable items
            it_odds = SortableTableWidgetItem(f"{odds:.2f}", odds)
            it_odds.setData(Qt.ItemDataRole.UserRole, (_live_cnt, _prem_cnt))
            it_live = SortableTableWidgetItem(fmt_wr(live_wr), live_wr)
            it_prem = SortableTableWidgetItem(fmt_wr(prem_wr), prem_wr)
            it_ev = SortableTableWidgetItem(ev_str, ev_val)
            
            cells = [it_odds, it_live, it_prem, it_ev]
            for c, it in enumerate(cells):
                it.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                self.table.setItem(r,c,it)
            color = neu if wr is None else pos if ev_val is not None and ev_val>0 else neg
            for c in range(4): self.table.item(r,c).setForeground(color)
        self.table.setSortingEnabled(True)
        
        # Re-apply sort to match the indicator
        header = self.table.horizontalHeader()
        self.table.sortItems(header.sortIndicatorSection(), header.sortIndicatorOrder())

    def update_ev_only(self):
        self.table.setSortingEnabled(False)
        bt = self.bettype_combo.currentText()
        # Fetch theme colors (same logic as fill_table)
        app = QApplication.instance()
        from PyQt6.QtGui import QColor as _QColor
        pos = app.property("positiveColor") if app else None
        neg = app.property("negativeColor") if app else None
        neu = app.property("neutralColor") if app else None
        if not isinstance(pos, _QColor): pos = _QColor('green')
        if not isinstance(neg, _QColor): neg = _QColor('red')
        if not isinstance(neu, _QColor): neu = _QColor('gray')
        for r in range(self.table.rowCount()):
            it_odds = self.table.item(r,0)
            odds = float(it_odds.text())
            counts = it_odds.data(Qt.ItemDataRole.UserRole)
            live_cnt, prem_cnt = counts if counts else (None, None)

            def parse_cell(s:str):
                if s in ("","N/A"): return None
                return float(s.rstrip('%'))/100.0
            live_wr = parse_cell(self.table.item(r,1).text())
            prem_wr = parse_cell(self.table.item(r,2).text())
            wr = live_wr if bt=="Live" else prem_wr
            cnt = live_cnt if bt=="Live" else prem_cnt

            ev_str, ev_val = fmt_ev(wr, odds, cnt)
            
            it = self.table.item(r,3)
            it.setText(ev_str)
            if isinstance(it, SortableTableWidgetItem):
                it.sort_value = ev_val
            
            color = neu if wr is None else pos if ev_val is not None and ev_val>0 else neg
            for c in range(4): self.table.item(r,c).setForeground(color)
        self.table.setSortingEnabled(True)
        
        # Re-apply sort
        header = self.table.horizontalHeader()
        self.table.sortItems(header.sortIndicatorSection(), header.sortIndicatorOrder())
        
        self.recompute_comparison_inline()

    # Comparison
    def render_compare(self, odds_a, odds_b, wr_a, wr_b, ev_a, ev_b, sport, bet_type, count_a, count_b):
        label_type = 'Live' if bet_type=='Live' else 'Prematch'
        if wr_a is None and wr_b is None:
            self.compare_result.setText(f"Both odds are missing {label_type} WR data: {odds_a:.2f} and {odds_b:.2f}."); return
        if wr_a is None or wr_b is None:
            present_odds = odds_b if wr_a is None else odds_a
            present_wr = wr_b if wr_a is None else wr_a
            present_ev = ev_b if wr_a is None else ev_a
            present_count = count_b if wr_a is None else count_a
            missing_odds = odds_a if wr_a is None else odds_b
            cnt = 'N/A' if present_count is None else str(present_count)
            self.compare_result.setText(
                f"Odds {missing_odds:.2f} is missing {label_type} WR.\n\n"
                f"Bet {present_odds:.2f} EV: {present_ev*100:.2f}% (WR: {present_wr*100:.2f}%, Count: {cnt})")
            return
        # Compare EVs directly
        if ev_a is not None and ev_b is not None:
            if ev_a > ev_b: better=f"Bet {odds_a:.2f} is better"
            elif ev_b > ev_a: better=f"Bet {odds_b:.2f} is better"
            else: better="Both bets are equal"
        else:
            better="Cannot compare (missing data)"
        ca = 'N/A' if count_a is None else str(count_a); cb = 'N/A' if count_b is None else str(count_b)
        txt=(f"{better} ({bet_type} - {sport})\n\n"+
             f"Bet {odds_a:.2f} EV: {ev_a*100:.2f}% (WR: {wr_a*100:.2f}%, Count: {ca})\n"+
             f"Bet {odds_b:.2f} EV: {ev_b*100:.2f}% (WR: {wr_b*100:.2f}%, Count: {cb})")
        notes=[]
        if count_a==1: notes.append(f"Note: Odds {odds_a:.2f} has only 1 count.")
        if count_b==1: notes.append(f"Note: Odds {odds_b:.2f} has only 1 count.")
        if notes: txt += "\n\n"+" ".join(notes)
        self.compare_result.setText(txt)

    def recompute_comparison_inline(self):
        try:
            odds_a = float(self.entry_odds_a.text().strip()); odds_b = float(self.entry_odds_b.text().strip())
        except ValueError:
            self.compare_result.setText(""); return
        bet_type = self.bettype_combo.currentText()
        k_a, k_b = round_odds_key(odds_a), round_odds_key(odds_b)
        miss_a = k_a not in self.odds_index; miss_b = k_b not in self.odds_index
        if miss_a and miss_b:
            self.compare_result.setText(f"Both odds not found in data: {odds_a:.2f} and {odds_b:.2f}."); return
        if miss_a or miss_b:
            pk = k_b if miss_a else k_a; po = odds_b if miss_a else odds_a
            wr_live, wr_pre, live_cnt, prem_cnt = self.odds_index[pk]
            wr_p = wr_live if bet_type=="Live" else wr_pre
            cnt_p = live_cnt if bet_type=="Live" else prem_cnt
            if wr_p is None:
                self.compare_result.setText((f"Odd {odds_a:.2f} not found. " if miss_a else f"Odd {odds_b:.2f} not found. ")+f"Odds {po:.2f} is in data but missing {bet_type} WR."); return
            ev_p = calc_ev(po, wr_p, cnt_p)
            if ev_p is None: ev_p = 0.0
            cnt_s='N/A' if cnt_p is None else str(cnt_p)
            missing = f"Odd {odds_a:.2f} not found." if miss_a else f"Odd {odds_b:.2f} not found."
            self.compare_result.setText(f"{missing}\n\nBet {po:.2f} EV: {ev_p*100:.2f}% (WR: {wr_p*100:.2f}%, Count: {cnt_s})")
            return
        wr_a_live, wr_a_pre, live_cnt_a, prem_cnt_a = self.odds_index[k_a]; wr_b_live, wr_b_pre, live_cnt_b, prem_cnt_b = self.odds_index[k_b]
        wr_a = wr_a_live if bet_type=="Live" else wr_a_pre; wr_b = wr_b_live if bet_type=="Live" else wr_b_pre
        cnt_a = live_cnt_a if bet_type=="Live" else prem_cnt_a
        cnt_b = live_cnt_b if bet_type=="Live" else prem_cnt_b
        ev_a = calc_ev(odds_a, wr_a, cnt_a)
        ev_b = calc_ev(odds_b, wr_b, cnt_b)
        self.render_compare(odds_a, odds_b, wr_a, wr_b, ev_a, ev_b, self.sport_combo.currentText(), bet_type, cnt_a, cnt_b)

    def compare_by_odds(self): self.recompute_comparison_inline()
    def on_bet_type_change(self): self.update_ev_only()

    def on_sport_change(self):
        sport = self.sport_combo.currentText()
        self._populate_tournament_combo(sport)
        self._populate_team_combo(sport)
        self._populate_provider_combo(sport)
        self.refresh_data(False)

    def on_tournament_change(self):
        self._apply_filters()

    def on_team_change(self):
        self._apply_filters()

    def on_provider_change(self):
        self._apply_filters()

    def _get_tournaments_for_sport(self, sport: str) -> List[str]:
        """Return normalized tournament names for the given sport, sorted by bet count descending."""
        counts: Dict[str, int] = {}
        for s, tournament, _, _, _, _, _, _ in self.matchbet_data:
            if s != sport:
                continue
            for label in tournament_filter_labels(tournament):
                counts[label] = counts.get(label, 0) + 1
        return sorted(counts.keys(), key=lambda t: counts[t], reverse=True)

    def _get_teams_for_sport(self, sport: str) -> List[str]:
        """Return team names for the given sport, sorted alphabetically."""
        teams: set = set()
        for s, _, matchup, bet, _, _, _, _ in self.matchbet_data:
            if s != sport:
                continue
            team = extract_team_from_bet(matchup, bet)
            if team:
                teams.add(team)
        return sorted(teams, key=str.lower)

    def _get_providers_for_sport(self, sport: str) -> List[str]:
        """Return provider names for the given sport, sorted alphabetically."""
        providers: set = set()
        for s, _, _, _, _, _, _, provider in self.matchbet_data:
            if s != sport:
                continue
            provider_name = str(provider or "").strip()
            if provider_name:
                providers.add(provider_name)
        return sorted(providers, key=str.lower)

    def _populate_tournament_combo(self, sport: str):
        """Repopulate the tournament combo for the given sport."""
        self.tournament_combo.blockSignals(True)
        self.tournament_combo.clear()
        self.tournament_combo.addItem("All Tournaments")
        for t in self._get_tournaments_for_sport(sport):
            self.tournament_combo.addItem(t)
        self.tournament_combo.setCurrentIndex(0)
        self.tournament_combo.blockSignals(False)

    def _populate_team_combo(self, sport: str):
        """Repopulate the team combo for the given sport."""
        self.team_combo.blockSignals(True)
        self.team_combo.clear()
        self.team_combo.addItem("All Teams")
        for t in self._get_teams_for_sport(sport):
            self.team_combo.addItem(t)
        self.team_combo.setCurrentIndex(0)
        self.team_combo.blockSignals(False)

    def _populate_provider_combo(self, sport: str):
        """Repopulate the provider combo for the given sport."""
        self.provider_combo.blockSignals(True)
        self.provider_combo.clear()
        self.provider_combo.addItem("All Providers")
        for provider in self._get_providers_for_sport(sport):
            self.provider_combo.addItem(provider)
        self.provider_combo.setCurrentIndex(0)
        self.provider_combo.blockSignals(False)

    def _apply_filters(self):
        """Apply all active filters (tournament + team) and update the table and comparison."""
        sport = self.sport_combo.currentText()
        tournament = self.tournament_combo.currentText()
        team = self.team_combo.currentText()
        provider = self.provider_combo.currentText()
        if not sport or sport not in self.data_cache:
            return

        all_tournaments = (tournament == "All Tournaments" or not tournament)
        all_teams = (team == "All Teams" or not team)
        all_providers = (provider == "All Providers" or not provider)

        if all_tournaments and all_teams and all_providers:
            # No filters — use full sport cache
            entry = self.data_cache[sport]
        else:
            # Filter matchbet_data then rebuild cache
            filtered: List[MatchBetTuple] = []
            for row in self.matchbet_data:
                if row[0] != sport:
                    continue
                # Tournament filter
                if not all_tournaments:
                    labels = tournament_filter_labels(row[1])
                    if tournament not in labels:
                        continue
                # Team filter — only include bets explicitly placed on the selected team
                if not all_teams:
                    bet_team = extract_team_from_bet(row[2], row[3])
                    if bet_team != team:
                        continue
                if not all_providers:
                    row_provider = str(row[7] or "").strip()
                    if row_provider != provider:
                        continue
                filtered.append(row)
            if not filtered:
                self.current_rows = []
                self.odds_index = {}
                self.fill_table([], self.bettype_combo.currentText())
                self.recompute_comparison_inline()
                return
            cache = process_bets_to_cache(filtered)
            entry = cache.get(sport)
            if entry is None:
                self.current_rows = []
                self.odds_index = {}
                self.fill_table([], self.bettype_combo.currentText())
                self.recompute_comparison_inline()
                return

        self.current_rows = entry.rows
        self.odds_index = entry.index
        self.fill_table(self.current_rows, self.bettype_combo.currentText())
        self.recompute_comparison_inline()

    def on_theme_toggled(self, checked: bool):
        self.dark_mode = checked
        theme_manager.apply_theme(QApplication.instance(), dark=checked)

    def open_settings_dialog(self):
        dlg = SettingsDialog(self)
        dlg.exec()

    def _update_nav_buttons(self, active):
        """Highlight only the active navigation button."""
        for btn in self._nav_buttons:
            btn.setChecked(btn is active)

    def show_statistics_panel(self, force_refresh_network: bool = False):
        """Switch to statistics panel view."""
        self.current_view = "statistics"
        self._update_nav_buttons(self.btn_statistics)
        self.table.hide()
        self.add_bet_panel.hide()
        self.pending_bets_panel.hide()
        self.statistics_panel.show()

        if getattr(self, "_statistics_panel_built", False):
            self._update_statistics_panel_content(force_refresh_network=force_refresh_network)
            return

        filters = self._get_statistics_filter_state()

        layout = self.statistics_panel.layout()
        if layout is None:
            layout = QVBoxLayout(self.statistics_panel)
            self.statistics_panel.setLayout(layout)

        self._clear_layout(layout)

        records = self._load_statistics_records(force_refresh_network=force_refresh_network)
        filtered_records, available_tournaments = self._filter_statistics_records(records, filters)
        snapshot = self._build_statistics_snapshot(filtered_records)
        self._statistics_snapshot = snapshot

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.Shape.NoFrame)
        scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)

        content = QWidget()
        content_layout = QVBoxLayout(content)
        content_layout.setContentsMargins(8, 8, 8, 12)
        content_layout.setSpacing(14)
        scroll.setWidget(content)
        layout.addWidget(scroll, 1)

        header_row = QHBoxLayout()
        header_row.setSpacing(12)

        title_block = QVBoxLayout()
        title = QLabel("Statistics Dashboard")
        title.setStyleSheet("font-size: 22px; font-weight: 700;")
        subtitle = QLabel("Scan performance by sport, tournament, status, and market type.")
        subtitle.setWordWrap(True)
        subtitle.setStyleSheet("font-size: 12px; color: #8b90a0;")
        title_block.addWidget(title)
        title_block.addWidget(subtitle)
        header_row.addLayout(title_block, 1)

        self.btn_stats_refresh = QPushButton("Refresh Stats")
        self.btn_stats_refresh.clicked.connect(lambda checked=False: self.show_statistics_panel(force_refresh_network=True))
        header_row.addWidget(self.btn_stats_refresh, 0, Qt.AlignmentFlag.AlignRight)
        content_layout.addLayout(header_row)

        filter_group = QGroupBox("Filters")
        filter_layout = QGridLayout(filter_group)
        filter_layout.setHorizontalSpacing(10)
        filter_layout.setVerticalSpacing(8)

        self.stats_sport_filter = QComboBox()
        self.stats_sport_filter.addItem("All Sports")
        sport_options = sorted({self._normalize_stats_sport(str(record[1] or "")) for record in records if str(record[1] or "").strip()}, key=str.lower)
        self.stats_sport_filter.addItems(sport_options)

        self.stats_scope_filter = QComboBox()
        self.stats_scope_filter.addItems(["All Bets", "Settled Bets", "Pending Bets"])

        self.stats_market_filter = QComboBox()
        self.stats_market_filter.addItems(["All Markets", "LIVE Only", "NOT LIVE Only"])

        self.stats_tournament_filter = QComboBox()
        self.stats_tournament_filter.addItem("All Tournaments")
        self.stats_tournament_filter.addItems(available_tournaments)

        self.stats_from_date_filter = QDateEdit()
        self.stats_from_date_filter.setCalendarPopup(True)
        self.stats_from_date_filter.setDisplayFormat("yyyy-MM-dd")
        self.stats_from_date_filter.setMinimumDate(QDate(2000, 1, 1))
        self.stats_from_date_filter.setMaximumDate(QDate.currentDate())
        if filters["from_date"]:
            from_date = QDate.fromString(filters["from_date"], "yyyy-MM-dd")
            if from_date.isValid():
                self.stats_from_date_filter.setDate(from_date)
        else:
            self.stats_from_date_filter.setDate(self._get_statistics_default_from_date(records))

        self.stats_search_filter = QLineEdit()
        self.stats_search_filter.setPlaceholderText("Search sport, tournament, matchup, bet, or result")

        self.btn_stats_reset = QPushButton("Reset")
        self.btn_stats_reset.clicked.connect(self._reset_statistics_filters)

        sport_filter_index = self.stats_sport_filter.findText(filters["sport"])
        if sport_filter_index >= 0:
            self.stats_sport_filter.setCurrentIndex(sport_filter_index)

        scope_filter_index = self.stats_scope_filter.findText(filters["scope"])
        if scope_filter_index >= 0:
            self.stats_scope_filter.setCurrentIndex(scope_filter_index)

        market_filter_index = self.stats_market_filter.findText(filters["market"])
        if market_filter_index >= 0:
            self.stats_market_filter.setCurrentIndex(market_filter_index)

        tournament_filter_index = self.stats_tournament_filter.findText(filters["tournament"])
        if tournament_filter_index >= 0:
            self.stats_tournament_filter.setCurrentIndex(tournament_filter_index)

        self.stats_search_filter.setText(filters["search"])

        self.stats_sport_filter.currentTextChanged.connect(self._schedule_statistics_refresh)
        self.stats_scope_filter.currentTextChanged.connect(self._schedule_statistics_refresh)
        self.stats_market_filter.currentTextChanged.connect(self._schedule_statistics_refresh)
        self.stats_tournament_filter.currentTextChanged.connect(self._schedule_statistics_refresh)
        self.stats_from_date_filter.dateChanged.connect(self._schedule_statistics_refresh)
        self.stats_search_filter.textChanged.connect(self._schedule_statistics_refresh)

        filter_layout.addWidget(QLabel("Sport"), 0, 0)
        filter_layout.addWidget(self.stats_sport_filter, 0, 1)
        filter_layout.addWidget(QLabel("Scope"), 0, 2)
        filter_layout.addWidget(self.stats_scope_filter, 0, 3)
        filter_layout.addWidget(QLabel("Market"), 0, 4)
        filter_layout.addWidget(self.stats_market_filter, 0, 5)
        filter_layout.addWidget(QLabel("Tournament"), 1, 0)
        filter_layout.addWidget(self.stats_tournament_filter, 1, 1)
        filter_layout.addWidget(QLabel("From date"), 1, 2)
        filter_layout.addWidget(self.stats_from_date_filter, 1, 3)
        filter_layout.addWidget(self.stats_search_filter, 1, 4, 1, 2)
        filter_layout.addWidget(self.btn_stats_reset, 0, 6, 2, 1)
        for column in (1, 3, 5):
            filter_layout.setColumnStretch(column, 1)
        filter_layout.setColumnStretch(4, 2)
        content_layout.addWidget(filter_group)

        overall = snapshot["overall"]
        card_grid = QGridLayout()
        card_grid.setHorizontalSpacing(10)
        card_grid.setVerticalSpacing(10)
        cards = [
            self._make_stat_card("Total Bets", str(overall["total_bets"]), "All records in the current filter set", "neutral"),
            self._make_stat_card("Settled / Pending", f'{overall["settled_bets"]} / {overall["pending_bets"]}', "Closed versus open bets", "neutral"),
            self._make_stat_card("Win Rate", self._format_percent(overall["win_rate"]), f'{overall["wins"]} wins, {overall["losses"]} losses', "positive"),
            self._make_stat_card("Profit", self._format_money(overall["profit"]), "Sum of settled profits", "positive" if overall["profit"] >= 0 else "negative"),
            self._make_stat_card("Average Odds", self._format_decimal(overall["avg_odds"]), "Mean odds across filtered bets", "neutral"),
            self._make_stat_card("LIVE / NOT LIVE", f'{overall["live_bets"]} / {overall["not_live_bets"]}', "Market split for the current view", "neutral"),
        ]
        for index, card in enumerate(cards):
            card_grid.addWidget(card, index // 3, index % 3)
        for column in range(3):
            card_grid.setColumnStretch(column, 1)
        content_layout.addLayout(card_grid)

        body_layout = QVBoxLayout()
        body_layout.setSpacing(12)

        left_group = QGroupBox("Sports and Tournaments")
        left_layout = QVBoxLayout(left_group)
        self.stats_tree = QTreeWidget()
        self.stats_tree.setAlternatingRowColors(True)
        self.stats_tree.setHeaderLabels(["Category", "Bets", "Settled", "Win Rate", "Profit"])
        tree_header = self.stats_tree.header()
        tree_header.setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        for column in (1, 2, 3, 4):
            tree_header.setSectionResizeMode(column, QHeaderView.ResizeMode.Fixed)
        self.stats_tree.setColumnWidth(1, 80)
        self.stats_tree.setColumnWidth(2, 80)
        self.stats_tree.setColumnWidth(3, 90)
        self.stats_tree.setColumnWidth(4, 95)
        self.stats_tree.setRootIsDecorated(True)
        self.stats_tree.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        self.stats_tree.itemSelectionChanged.connect(self._update_statistics_detail)

        sports = sorted(snapshot["sports"].items(), key=lambda item: item[1]["total"], reverse=True)
        for sport, sport_data in sports:
            sport_item = QTreeWidgetItem(self.stats_tree)
            sport_item.setText(0, sport)
            sport_item.setText(1, str(sport_data["total"]))
            sport_item.setText(2, str(sport_data["settled"]))
            sport_item.setText(3, self._format_percent(sport_data["win_rate"]))
            sport_item.setText(4, self._format_money(sport_data["profit"]))
            sport_item.setData(0, Qt.ItemDataRole.UserRole, ("sport", sport, ""))
            sport_item.setExpanded(True)
            for tournament, tournament_data in sorted(sport_data["tournaments"].items(), key=lambda item: item[1]["total"], reverse=True):
                tournament_item = QTreeWidgetItem(sport_item)
                tournament_item.setText(0, tournament)
                tournament_item.setText(1, str(tournament_data["total"]))
                tournament_item.setText(2, str(tournament_data["settled"]))
                tournament_item.setText(3, self._format_percent(tournament_data["win_rate"]))
                tournament_item.setText(4, self._format_money(tournament_data["profit"]))
                tournament_item.setData(0, Qt.ItemDataRole.UserRole, ("tournament", sport, tournament))

        if self.stats_tree.topLevelItemCount() > 0:
            self.stats_tree.setCurrentItem(self.stats_tree.topLevelItem(0))

        left_layout.addWidget(self.stats_tree)
        body_layout.addWidget(left_group)

        right_widget = QWidget()
        right_layout = QVBoxLayout(right_widget)
        right_layout.setContentsMargins(0, 0, 0, 0)
        right_layout.setSpacing(12)

        detail_group = QGroupBox("Selection Details")
        detail_layout = QVBoxLayout(detail_group)
        self.stats_detail_label = QLabel()
        self.stats_detail_label.setWordWrap(True)
        self.stats_detail_label.setTextInteractionFlags(Qt.TextInteractionFlag.TextSelectableByMouse)
        detail_layout.addWidget(self.stats_detail_label)
        right_layout.addWidget(detail_group)

        insight_group = QGroupBox("Quick Insights")
        insight_layout = QVBoxLayout(insight_group)
        self.stats_insight_label = QLabel(self._build_statistics_insights(snapshot))
        self.stats_insight_label.setWordWrap(True)
        self.stats_insight_label.setTextInteractionFlags(Qt.TextInteractionFlag.TextSelectableByMouse)
        insight_layout.addWidget(self.stats_insight_label)
        right_layout.addWidget(insight_group)

        roadmap_group = QGroupBox("Ideas to Add Next")
        roadmap_layout = QVBoxLayout(roadmap_group)
        self.stats_roadmap_label = QLabel(
            "<ul>"
            "<li>Win-rate and profit trend charts over time.</li>"
            "<li>Streak tracking for current hot and cold runs.</li>"
            "<li>A search box that filters down to a single bet or matchup.</li>"
            "<li>Top and bottom performer chips for quick navigation.</li>"
            "<li>Export buttons for CSV or image snapshots of the dashboard.</li>"
            "</ul>"
        )
        self.stats_roadmap_label.setWordWrap(True)
        roadmap_layout.addWidget(self.stats_roadmap_label)
        right_layout.addWidget(roadmap_group)

        body_layout.addWidget(right_widget)
        content_layout.addLayout(body_layout, 1)

        self._statistics_panel_built = True
        self._update_statistics_detail()

    def _clear_layout(self, layout):
        while layout.count():
            item = layout.takeAt(0)
            widget = item.widget()
            if widget is not None:
                widget.deleteLater()
            child_layout = item.layout()
            if child_layout is not None:
                self._clear_layout(child_layout)

    def _normalize_stats_sport(self, sport: str) -> str:
        sport_name = str(sport or "").strip()
        if not sport_name:
            return "Unknown"
        if sport_name == "CStwo":
            return "CS2"
        return sport_name

    def _get_statistics_filter_state(self) -> Dict[str, str]:
        return {
            "sport": self.stats_sport_filter.currentText() if hasattr(self, "stats_sport_filter") else "All Sports",
            "tournament": self.stats_tournament_filter.currentText() if hasattr(self, "stats_tournament_filter") else "All Tournaments",
            "scope": self.stats_scope_filter.currentText() if hasattr(self, "stats_scope_filter") else "All Bets",
            "market": self.stats_market_filter.currentText() if hasattr(self, "stats_market_filter") else "All Markets",
            "from_date": self.stats_from_date_filter.date().toString("yyyy-MM-dd") if hasattr(self, "stats_from_date_filter") else "",
            "search": self.stats_search_filter.text().strip() if hasattr(self, "stats_search_filter") else "",
        }

    def _get_statistics_default_from_date(self, records: List[Tuple]) -> QDate:
        dates = []
        for record in records:
            if len(record) < 12:
                continue
            parsed = self._parse_statistics_datetime(record[10])
            if parsed is not None:
                dates.append(parsed.date())
        if dates:
            oldest = min(dates)
            return QDate(oldest.year, oldest.month, oldest.day)
        return QDate.currentDate()

    def _parse_statistics_datetime(self, value) -> Optional[datetime]:
        if not value:
            return None
        try:
            text = str(value).replace("Z", "+00:00")
            return datetime.fromisoformat(text)
        except Exception:
            return None

    def _load_statistics_records(self, force_refresh_network: bool = False) -> List[Tuple]:
        try:
            return self.db.fetch_all_bets(force_refresh=force_refresh_network)
        except Exception as e:
            print(f"Error fetching statistics data: {e}")
            return []

    def _update_statistics_panel_content(self, force_refresh_network: bool = False):
        if not hasattr(self, "stats_tree"):
            self.show_statistics_panel(force_refresh_network=force_refresh_network)
            return

        records = self._load_statistics_records(force_refresh_network=force_refresh_network)
        filters = self._get_statistics_filter_state()
        filtered_records, available_tournaments = self._filter_statistics_records(records, filters)
        snapshot = self._build_statistics_snapshot(filtered_records)
        self._statistics_snapshot = snapshot

        self.stats_tournament_filter.blockSignals(True)
        current_tournament = self.stats_tournament_filter.currentText()
        self.stats_tournament_filter.clear()
        self.stats_tournament_filter.addItem("All Tournaments")
        self.stats_tournament_filter.addItems(available_tournaments)
        if current_tournament and self.stats_tournament_filter.findText(current_tournament) >= 0:
            self.stats_tournament_filter.setCurrentText(current_tournament)
        else:
            self.stats_tournament_filter.setCurrentIndex(0)
        self.stats_tournament_filter.blockSignals(False)

        self._populate_statistics_tree(snapshot)

        overall = snapshot["overall"]
        if hasattr(self, "stats_detail_label"):
            self.stats_detail_label.setText("Select a sport or tournament to inspect the filtered data.")
        if hasattr(self, "stats_insight_label"):
            self.stats_insight_label.setText(self._build_statistics_insights(snapshot))

        self.stats_tree.setEnabled(True)
        if self.stats_tree.topLevelItemCount() > 0:
            self.stats_tree.setCurrentItem(self.stats_tree.topLevelItem(0))
        self._update_statistics_detail()

    def _populate_statistics_tree(self, snapshot: Dict[str, dict]):
        self.stats_tree.blockSignals(True)
        self.stats_tree.clear()

        sports = sorted(snapshot["sports"].items(), key=lambda item: item[1]["total"], reverse=True)
        for sport, sport_data in sports:
            sport_item = QTreeWidgetItem(self.stats_tree)
            sport_item.setText(0, sport)
            sport_item.setText(1, str(sport_data["total"]))
            sport_item.setText(2, str(sport_data["settled"]))
            sport_item.setText(3, self._format_percent(sport_data["win_rate"]))
            sport_item.setText(4, self._format_money(sport_data["profit"]))
            sport_item.setData(0, Qt.ItemDataRole.UserRole, ("sport", sport, ""))
            sport_item.setExpanded(True)
            for tournament, tournament_data in sorted(sport_data["tournaments"].items(), key=lambda item: item[1]["total"], reverse=True):
                tournament_item = QTreeWidgetItem(sport_item)
                tournament_item.setText(0, tournament)
                tournament_item.setText(1, str(tournament_data["total"]))
                tournament_item.setText(2, str(tournament_data["settled"]))
                tournament_item.setText(3, self._format_percent(tournament_data["win_rate"]))
                tournament_item.setText(4, self._format_money(tournament_data["profit"]))
                tournament_item.setData(0, Qt.ItemDataRole.UserRole, ("tournament", sport, tournament))

        if self.stats_tree.topLevelItemCount() > 0:
            self.stats_tree.setCurrentItem(self.stats_tree.topLevelItem(0))
        self.stats_tree.blockSignals(False)

    def _filter_statistics_records(self, records: List[Tuple], filters: Dict[str, str]) -> Tuple[List[dict], List[str]]:
        base_records: List[dict] = []
        tournaments = set()

        for record in records:
            if len(record) < 12:
                continue

            _, sport, tournament, matchup, bet, live_status, odds, bet_amount, result, profit, date_created, _provider = record
            sport_name = self._normalize_stats_sport(sport)
            tournament_name = normalize_tournament_name(str(tournament or "")) or "Unspecified"
            result_text = str(result or "").strip()
            live_status_text = str(live_status or "").strip()
            is_live = "LIVE" in live_status_text.upper() and "NOT" not in live_status_text.upper()
            is_settled = bool(result_text)
            parsed_date = self._parse_statistics_datetime(date_created)

            if filters["sport"] != "All Sports" and sport_name != filters["sport"]:
                continue
            if filters["scope"] == "Settled Bets" and not is_settled:
                continue
            if filters["scope"] == "Pending Bets" and is_settled:
                continue
            if filters["market"] == "LIVE Only" and not is_live:
                continue
            if filters["market"] == "NOT LIVE Only" and is_live:
                continue
            if filters.get("from_date"):
                from_date = QDate.fromString(filters["from_date"], "yyyy-MM-dd")
                if from_date.isValid() and parsed_date is not None:
                    record_date = QDate(parsed_date.year, parsed_date.month, parsed_date.day)
                    if record_date < from_date:
                        continue
                elif from_date.isValid() and parsed_date is None:
                    continue

            parsed = {
                "id": record[0],
                "sport": sport_name,
                "tournament": tournament_name,
                "matchup": str(matchup or ""),
                "bet": str(bet or ""),
                "live_status": live_status_text,
                "odds": float(odds or 0.0),
                "bet_amount": float(bet_amount or 0.0),
                "result": result_text,
                "profit": float(profit or 0.0),
                "date_created": str(date_created or ""),
                "parsed_date": parsed_date,
                "is_live": is_live,
                "is_settled": is_settled,
                "is_win": result_text.lower() == "win",
                "is_loss": result_text.lower() == "lose",
            }

            base_records.append(parsed)
            tournaments.add(tournament_name)

        final_records: List[dict] = []
        search_text = filters["search"].lower()
        for parsed in base_records:
            if filters["tournament"] != "All Tournaments" and parsed["tournament"] != filters["tournament"]:
                continue
            if search_text:
                haystack = " ".join([
                    parsed["sport"], parsed["tournament"], parsed["matchup"],
                    parsed["bet"], parsed["live_status"], parsed["result"]
                ]).lower()
                if search_text not in haystack:
                    continue
            final_records.append(parsed)

        return final_records, sorted(tournaments, key=str.lower)

    def _build_statistics_snapshot(self, records: List[dict]) -> Dict[str, dict]:
        snapshot: Dict[str, dict] = {
            "records": records,
            "sports": {},
            "overall": {
                "total_bets": 0,
                "settled_bets": 0,
                "pending_bets": 0,
                "wins": 0,
                "losses": 0,
                "profit": 0.0,
                "odds_sum": 0.0,
                "live_bets": 0,
                "not_live_bets": 0,
                "live_settled": 0,
                "live_wins": 0,
                "not_live_settled": 0,
                "not_live_wins": 0,
            },
        }

        for record in records:
            overall = snapshot["overall"]
            sport = record["sport"]
            tournament = record["tournament"]
            sport_bucket = snapshot["sports"].setdefault(
                sport,
                {
                    "total": 0,
                    "settled": 0,
                    "pending": 0,
                    "wins": 0,
                    "losses": 0,
                    "profit": 0.0,
                    "odds_sum": 0.0,
                    "live_total": 0,
                    "live_settled": 0,
                    "live_wins": 0,
                    "not_live_total": 0,
                    "not_live_settled": 0,
                    "not_live_wins": 0,
                    "tournaments": {},
                },
            )

            tournament_bucket = sport_bucket["tournaments"].setdefault(
                tournament,
                {
                    "total": 0,
                    "settled": 0,
                    "pending": 0,
                    "wins": 0,
                    "losses": 0,
                    "profit": 0.0,
                    "odds_sum": 0.0,
                    "live_total": 0,
                    "live_settled": 0,
                    "live_wins": 0,
                    "not_live_total": 0,
                    "not_live_settled": 0,
                    "not_live_wins": 0,
                },
            )

            for bucket in (sport_bucket, tournament_bucket):
                bucket["total"] += 1
                bucket["profit"] += record["profit"]
                bucket["odds_sum"] += record["odds"]
                if record["is_settled"]:
                    bucket["settled"] += 1
                    if record["is_win"]:
                        bucket["wins"] += 1
                    elif record["is_loss"]:
                        bucket["losses"] += 1
                else:
                    bucket["pending"] += 1
                if record["is_live"]:
                    bucket["live_total"] += 1
                    if record["is_settled"]:
                        bucket["live_settled"] += 1
                        if record["is_win"]:
                            bucket["live_wins"] += 1
                else:
                    bucket["not_live_total"] += 1
                    if record["is_settled"]:
                        bucket["not_live_settled"] += 1
                        if record["is_win"]:
                            bucket["not_live_wins"] += 1

            overall["total_bets"] += 1
            overall["profit"] += record["profit"]
            overall["odds_sum"] += record["odds"]
            if record["is_live"]:
                overall["live_bets"] += 1
            else:
                overall["not_live_bets"] += 1

            if record["is_settled"]:
                overall["settled_bets"] += 1
                if record["is_win"]:
                    overall["wins"] += 1
                elif record["is_loss"]:
                    overall["losses"] += 1
                if record["is_live"]:
                    overall["live_settled"] += 1
                    if record["is_win"]:
                        overall["live_wins"] += 1
                else:
                    overall["not_live_settled"] += 1
                    if record["is_win"]:
                        overall["not_live_wins"] += 1
            else:
                overall["pending_bets"] += 1

        overall["avg_odds"] = overall["odds_sum"] / overall["total_bets"] if overall["total_bets"] > 0 else None
        overall["win_rate"] = overall["wins"] / overall["settled_bets"] if overall["settled_bets"] > 0 else None
        overall["live_win_rate"] = overall["live_wins"] / overall["live_settled"] if overall["live_settled"] > 0 else None
        overall["not_live_win_rate"] = overall["not_live_wins"] / overall["not_live_settled"] if overall["not_live_settled"] > 0 else None

        for sport_data in snapshot["sports"].values():
            sport_data["win_rate"] = sport_data["wins"] / sport_data["settled"] if sport_data["settled"] > 0 else None
            sport_data["avg_odds"] = sport_data["odds_sum"] / sport_data["total"] if sport_data["total"] > 0 else None
            for tournament_data in sport_data["tournaments"].values():
                tournament_data["win_rate"] = tournament_data["wins"] / tournament_data["settled"] if tournament_data["settled"] > 0 else None
                tournament_data["avg_odds"] = tournament_data["odds_sum"] / tournament_data["total"] if tournament_data["total"] > 0 else None

        snapshot["top_sports_by_volume"] = sorted(snapshot["sports"].items(), key=lambda item: item[1]["total"], reverse=True)
        snapshot["top_sports_by_profit"] = sorted(snapshot["sports"].items(), key=lambda item: item[1]["profit"], reverse=True)
        snapshot["top_tournaments_by_volume"] = self._flatten_tournaments(snapshot, sort_key="total")
        snapshot["top_tournaments_by_profit"] = self._flatten_tournaments(snapshot, sort_key="profit")
        snapshot["top_tournaments_by_win_rate"] = [
            item for item in self._flatten_tournaments(snapshot, sort_key="win_rate", reverse=True)
            if item[2]["settled"] > 0
        ]
        return snapshot

    def _flatten_tournaments(self, snapshot: Dict[str, dict], sort_key: str = "total", reverse: bool = True):
        items = []
        for sport, sport_data in snapshot["sports"].items():
            for tournament, tournament_data in sport_data["tournaments"].items():
                items.append((sport, tournament, tournament_data))
        return sorted(items, key=lambda item: item[2][sort_key] if item[2][sort_key] is not None else float("-inf"), reverse=reverse)

    def _format_percent(self, value: Optional[float]) -> str:
        return "N/A" if value is None else f"{value * 100:.1f}%"

    def _format_decimal(self, value: Optional[float]) -> str:
        return "N/A" if value is None else f"{value:.2f}"

    def _format_money(self, value: float) -> str:
        sign = "+" if value >= 0 else ""
        return f"{sign}{value:.2f}"

    def _make_stat_card(self, title: str, value: str, subtitle: str, tone: str = "neutral") -> QWidget:
        app = QApplication.instance()
        positive_color = app.property("positiveColor") if app else None
        negative_color = app.property("negativeColor") if app else None
        neutral_color = app.property("neutralColor") if app else None
        if not isinstance(positive_color, QColor):
            positive_color = QColor("#16a34a")
        if not isinstance(negative_color, QColor):
            negative_color = QColor("#dc2626")
        if not isinstance(neutral_color, QColor):
            neutral_color = QColor("#6b7280")

        tone_color = neutral_color
        if tone == "positive":
            tone_color = positive_color
        elif tone == "negative":
            tone_color = negative_color

        card = QFrame()
        card.setStyleSheet(
            "QFrame {"
            " background-color: rgba(255, 255, 255, 0.03);"
            f" border: 1px solid {tone_color.name()};"
            " border-radius: 14px;"
            " }"
        )
        card_layout = QVBoxLayout(card)
        card_layout.setContentsMargins(14, 12, 14, 12)
        card_layout.setSpacing(4)

        title_label = QLabel(title)
        title_label.setStyleSheet(f"font-size: 11px; font-weight: 700; color: {tone_color.name()};")
        value_label = QLabel(value)
        value_label.setStyleSheet("font-size: 20px; font-weight: 800;")
        subtitle_label = QLabel(subtitle)
        subtitle_label.setWordWrap(True)
        subtitle_label.setStyleSheet("font-size: 11px; color: #8b90a0;")

        card_layout.addWidget(title_label)
        card_layout.addWidget(value_label)
        card_layout.addWidget(subtitle_label)
        return card

    def _build_statistics_insights(self, snapshot: Dict[str, dict]) -> str:
        overall = snapshot["overall"]
        lines = ["<b>Current snapshot</b><br>"]
        lines.append(f'Total records: {overall["total_bets"]}<br>')
        lines.append(f'Settled bets: {overall["settled_bets"]}<br>')
        lines.append(f'Pending bets: {overall["pending_bets"]}<br>')
        lines.append(f'Win rate: {self._format_percent(overall["win_rate"])}<br>')
        lines.append(f'Profit: {self._format_money(overall["profit"])}<br><br>')

        top_sport = snapshot["top_sports_by_volume"][0] if snapshot["top_sports_by_volume"] else None
        if top_sport is not None:
            lines.append(f'<b>Most active sport:</b> {top_sport[0]} ({top_sport[1]["total"]} bets)<br>')

        best_profit_tourney = snapshot["top_tournaments_by_profit"][0] if snapshot["top_tournaments_by_profit"] else None
        if best_profit_tourney is not None:
            lines.append(
                f'<b>Best tournament by profit:</b> {best_profit_tourney[1]} / {best_profit_tourney[0]} '
                f'({self._format_money(best_profit_tourney[2]["profit"])})<br>'
            )

        best_wr_tourney = next((item for item in snapshot["top_tournaments_by_win_rate"] if item[2]["settled"] >= 3), None)
        if best_wr_tourney is not None:
            lines.append(
                f'<b>Best tournament by win rate:</b> {best_wr_tourney[1]} / {best_wr_tourney[0]} '
                f'({self._format_percent(best_wr_tourney[2]["win_rate"])})<br>'
            )

        live_wr = snapshot["overall"].get("live_win_rate")
        not_live_wr = snapshot["overall"].get("not_live_win_rate")
        lines.append(f'<br><b>LIVE win rate:</b> {self._format_percent(live_wr)}<br>')
        lines.append(f'<b>NOT LIVE win rate:</b> {self._format_percent(not_live_wr)}<br>')
        return "".join(lines)

    def _update_statistics_detail(self):
        tree = getattr(self, "stats_tree", None)
        detail_label = getattr(self, "stats_detail_label", None)
        if tree is None or detail_label is None:
            return

        current = tree.currentItem()
        snapshot = getattr(self, "_statistics_snapshot", {})
        if not current or not snapshot:
            detail_label.setText("Select a sport or tournament to inspect the filtered data.")
            return

        payload = current.data(0, Qt.ItemDataRole.UserRole)
        if not payload:
            detail_label.setText("Select a sport or tournament to inspect the filtered data.")
            return

        kind, sport, tournament = payload
        if kind == "sport":
            bucket = snapshot["sports"].get(sport)
            if bucket is None:
                return
            title = f"<b>{sport}</b>"
            subtitle = f'{bucket["total"]} bets, {bucket["settled"]} settled, {bucket["pending"]} pending'
            top_tournaments = sorted(bucket["tournaments"].items(), key=lambda item: item[1]["total"], reverse=True)[:3]
            extra_lines = "".join(
                f'<li>{name}: {data["total"]} bets, {self._format_percent(data["win_rate"])} win rate, {self._format_money(data["profit"])} profit</li>'
                for name, data in top_tournaments
            )
            detail_header = "Top related tournaments"
        else:
            bucket = snapshot["sports"].get(sport, {}).get("tournaments", {}).get(tournament)
            if bucket is None:
                return
            title = f"<b>{tournament}</b><br><span style='font-size: 12px;'>in {sport}</span>"
            subtitle = f'{bucket["total"]} bets, {bucket["settled"]} settled, {bucket["pending"]} pending'
            extra_lines = "".join([
                f'<li>LIVE split: {bucket["live_total"]} total, {bucket["live_settled"]} settled, '
                f'{self._format_percent(bucket["live_wins"] / bucket["live_settled"] if bucket["live_settled"] else None)} win rate</li>',
                f'<li>NOT LIVE split: {bucket["not_live_total"]} total, {bucket["not_live_settled"]} settled, '
                f'{self._format_percent(bucket["not_live_wins"] / bucket["not_live_settled"] if bucket["not_live_settled"] else None)} win rate</li>'
            ])
            detail_header = "Market split"

        detail_html = [
            title,
            f'<br><b>Summary:</b> {subtitle}<br>',
            f'<b>Win rate:</b> {self._format_percent(bucket["win_rate"])}<br>',
            f'<b>Profit:</b> {self._format_money(bucket["profit"])}<br>',
            f'<b>Average odds:</b> {self._format_decimal(bucket["avg_odds"])}<br>',
            f'<b>LIVE / NOT LIVE:</b> {bucket["live_total"]} / {bucket["not_live_total"]}<br><br>',
            f'<b>{detail_header}:</b><ul>',
            extra_lines,
            '</ul>'
        ]
        detail_label.setText("".join(detail_html))

    def _refresh_statistics_panel(self, *args):
        self._schedule_statistics_refresh(*args)

    def _schedule_statistics_refresh(self, *args):
        if self.current_view != "statistics":
            return
        if self._statistics_refresh_pending:
            return
        self._statistics_refresh_pending = True

        def _run_refresh():
            self._statistics_refresh_pending = False
            if self.current_view != "statistics":
                return
            if getattr(self, "_statistics_panel_built", False):
                self._update_statistics_panel_content()
            else:
                self.show_statistics_panel()

        QTimer.singleShot(0, _run_refresh)

    def _reset_statistics_filters(self):
        for widget, value in (
            (getattr(self, "stats_sport_filter", None), "All Sports"),
            (getattr(self, "stats_scope_filter", None), "All Bets"),
            (getattr(self, "stats_market_filter", None), "All Markets"),
            (getattr(self, "stats_tournament_filter", None), "All Tournaments"),
        ):
            if widget is not None:
                widget.blockSignals(True)
                widget.setCurrentText(value)
                widget.blockSignals(False)
        if hasattr(self, "stats_from_date_filter"):
            self.stats_from_date_filter.blockSignals(True)
            self.stats_from_date_filter.setDate(self._get_statistics_default_from_date(self._load_statistics_records()))
            self.stats_from_date_filter.blockSignals(False)
        if hasattr(self, "stats_search_filter"):
            self.stats_search_filter.blockSignals(True)
            self.stats_search_filter.clear()
            self.stats_search_filter.blockSignals(False)
        if getattr(self, "_statistics_panel_built", False):
            self._update_statistics_panel_content()
        else:
            self.show_statistics_panel()

    def show_data_table_panel(self):
        """Switch to data table view"""
        self.current_view = "table"
        self._update_nav_buttons(self.btn_data_table)
        self.statistics_panel.hide()
        self.add_bet_panel.hide()
        self.pending_bets_panel.hide()
        self.table.show()

    # ------------------------------------------------------------------
    # Add Bet Panel
    # ------------------------------------------------------------------
    def _build_add_bet_panel(self, parent_layout):
        """Create the Add Bet / Settle Bet form panel."""
        self.add_bet_panel = QWidget()
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        inner = QWidget()
        form_layout = QVBoxLayout(inner)
        form_layout.setContentsMargins(20, 20, 20, 20)
        form_layout.setSpacing(10)

        self.add_bet_title = QLabel("Add New Bet")
        self.add_bet_title.setStyleSheet("font-size: 18px; font-weight: bold;")
        form_layout.addWidget(self.add_bet_title)

        form = QFormLayout()
        form.setSpacing(8)

        # Sport
        self.form_sport = QComboBox()
        self.form_sport.setEditable(True)
        form.addRow("Sport:", self.form_sport)

        # Tournament
        self.form_tournament = QComboBox()
        self.form_tournament.setEditable(True)
        form.addRow("Tournament:", self.form_tournament)

        # Provider
        self.form_provider = QComboBox()
        self.form_provider.setEditable(True)
        form.addRow("Provider:", self.form_provider)

        # Matchup
        matchup_layout = QHBoxLayout()
        self.form_matchup_team_a = QComboBox()
        self.form_matchup_team_a.setEditable(True)
        self.form_matchup_team_b = QComboBox()
        self.form_matchup_team_b.setEditable(True)
        matchup_layout.addWidget(self.form_matchup_team_a, 1)
        # Center "vs" text
        lbl_vs = QLabel("vs")
        lbl_vs.setAlignment(Qt.AlignmentFlag.AlignCenter)
        matchup_layout.addWidget(lbl_vs)
        matchup_layout.addWidget(self.form_matchup_team_b, 1)
        form.addRow("Matchup:", matchup_layout)

        # Bet
        self.form_bet = QLineEdit()
        self.form_bet.setPlaceholderText("e.g. Match: Team A")
        form.addRow("Bet:", self.form_bet)

        # Live Status
        self.form_live_status = QComboBox()
        self.form_live_status.addItems(["LIVE", "NOT LIVE"])
        form.addRow("Live Status:", self.form_live_status)

        # Odds
        self.form_odds = QLineEdit()
        self.form_odds.setPlaceholderText("e.g. 1.75")
        form.addRow("Odds:", self.form_odds)

        # Bet Amount
        self.form_bet_amount = QLineEdit()
        self.form_bet_amount.setPlaceholderText("e.g. 10.00")
        form.addRow("Bet Amount:", self.form_bet_amount)

        # Result
        self.form_result = QComboBox()
        self.form_result.addItems(["(Pending)", "Win", "Lose"])
        form.addRow("Result:", self.form_result)

        # Profit (auto-calculated)
        self.form_profit = QLineEdit()
        self.form_profit.setPlaceholderText("Auto-calculated from odds & amount")
        self.form_profit.setReadOnly(True)
        form.addRow("Profit:", self.form_profit)

        form_layout.addLayout(form)

        # Buttons
        btn_row = QHBoxLayout()
        self.btn_save_bet = QPushButton("Save Bet")
        self.btn_save_bet.clicked.connect(self.save_bet)
        btn_row.addWidget(self.btn_save_bet)

        btn_clear = QPushButton("Clear Form")
        btn_clear.clicked.connect(self.clear_bet_form)
        btn_row.addWidget(btn_clear)

        btn_cancel = QPushButton("Cancel")
        
        def on_add_bet_cancel():
            if self.editing_bet_id:
                self._reset_form_to_add_mode()
                self.show_pending_bets_panel()
            else:
                self.show_data_table_panel()

        btn_cancel.clicked.connect(on_add_bet_cancel)
        btn_row.addWidget(btn_cancel)

        form_layout.addLayout(btn_row)
        form_layout.addStretch(1)

        scroll.setWidget(inner)
        panel_layout = QVBoxLayout(self.add_bet_panel)
        panel_layout.setContentsMargins(0, 0, 0, 0)
        panel_layout.addWidget(scroll)
        parent_layout.addWidget(self.add_bet_panel, 1)
        self.add_bet_panel.hide()

        # Connect signals for auto-profit calculation
        self.form_result.currentTextChanged.connect(self._auto_calc_profit)
        self.form_odds.textChanged.connect(self._auto_calc_profit)
        self.form_bet_amount.textChanged.connect(self._auto_calc_profit)

        # Connect signal for sport change to update tournament/team dropdowns
        self.form_sport.currentTextChanged.connect(self._on_form_sport_changed)

    def _on_form_sport_changed(self, sport: str):
        """Update tournament and team combo boxes when the sport changes."""
        self.form_tournament.blockSignals(True)
        self.form_provider.blockSignals(True)
        self.form_matchup_team_a.blockSignals(True)
        self.form_matchup_team_b.blockSignals(True)

        current_tournament = self.form_tournament.currentText()
        current_provider = self.form_provider.currentText()
        current_team_a = self.form_matchup_team_a.currentText()
        current_team_b = self.form_matchup_team_b.currentText()

        self.form_tournament.clear()
        self.form_provider.clear()
        self.form_matchup_team_a.clear()
        self.form_matchup_team_b.clear()

        if sport:
            tournaments = self._get_tournaments_for_sport(sport)
            providers = self._get_providers_for_sport(sport)
            teams = self._get_teams_for_sport(sport)
            if tournaments:
                self.form_tournament.addItems(tournaments)
            if providers:
                self.form_provider.addItems(providers)
            if teams:
                self.form_matchup_team_a.addItems(teams)
                self.form_matchup_team_b.addItems(teams)

        self.form_tournament.setCurrentText(current_tournament)
        self.form_provider.setCurrentText(current_provider)
        self.form_matchup_team_a.setCurrentText(current_team_a)
        self.form_matchup_team_b.setCurrentText(current_team_b)

        self.form_tournament.blockSignals(False)
        self.form_provider.blockSignals(False)
        self.form_matchup_team_a.blockSignals(False)
        self.form_matchup_team_b.blockSignals(False)

    def _auto_calc_profit(self):
        """Recalculate profit when result, odds, or bet amount change."""
        result = self.form_result.currentText()
        if result == "(Pending)":
            self.form_profit.clear()
            return
        try:
            odds = float(self.form_odds.text().strip().replace(",", "."))
            amount = float(self.form_bet_amount.text().strip().replace(",", "."))
        except (ValueError, AttributeError):
            self.form_profit.clear()
            return
        if result == "Win":
            profit = (odds - 1.0) * amount
        else:  # Lose
            profit = -amount
        self.form_profit.setText(f"{profit:.2f}")

    def _populate_form_sports(self):
        """Populate the sport combo in the add-bet form from existing data."""
        self.form_sport.blockSignals(True)
        current = self.form_sport.currentText()
        self.form_sport.clear()
        sports = sorted(set(
            s for s, *_ in self.matchbet_data
        )) if self.matchbet_data else []
        # Also include sports from db in case matchbet_data is empty
        try:
            db_sports = self.db.get_distinct_sports()
            sports = sorted(set(sports) | set(db_sports))
        except Exception:
            pass
        for s in sports:
            self.form_sport.addItem(s)
        if current:
            self.form_sport.setCurrentText(current)
        self.form_sport.blockSignals(False)

    def _populate_form_providers(self):
        """Populate the provider combo in the add-bet form from existing data."""
        self.form_provider.blockSignals(True)
        current = self.form_provider.currentText()
        self.form_provider.clear()
        try:
            providers = self.db.get_distinct_providers()
        except Exception:
            providers = []
        for provider in providers:
            self.form_provider.addItem(provider)
        if current:
            self.form_provider.setCurrentText(current)
        self.form_provider.blockSignals(False)

    def show_add_bet_panel(self):
        """Switch to add-bet form in add mode."""
        self.current_view = "add_bet"
        self._update_nav_buttons(self.btn_add_bet)
        self.table.hide()
        self.statistics_panel.hide()
        self.pending_bets_panel.hide()
        self.add_bet_panel.show()
        self._reset_form_to_add_mode()

    def _reset_form_to_add_mode(self):
        """Reset the form to add-bet mode (clear fields, enable all)."""
        self.editing_bet_id = None
        self.add_bet_title.setText("Add New Bet")
        self.btn_save_bet.setText("Save Bet")
        self._populate_form_sports()
        self._populate_form_providers()

        # Enable all fields
        for w in [self.form_sport, self.form_tournament, 
                  self.form_provider, self.form_matchup_team_a, self.form_matchup_team_b,
                  self.form_bet, self.form_live_status,
                  self.form_odds, self.form_bet_amount]:
            w.setEnabled(True)

        self.form_result.setEnabled(True)
        self.form_profit.setReadOnly(True)
        self.clear_bet_form()

    def clear_bet_form(self):
        """Clear all form fields to defaults."""
        if self.editing_bet_id is None:
            self.form_sport.setCurrentIndex(0) if self.form_sport.count() > 0 else self.form_sport.setCurrentText("")
        self.form_tournament.setCurrentText("")
        self.form_provider.setCurrentText("")
        self.form_matchup_team_a.setCurrentText("")
        self.form_matchup_team_b.setCurrentText("")
        self.form_bet.clear()
        self.form_live_status.setCurrentIndex(0)
        self.form_odds.clear()
        self.form_bet_amount.clear()
        self.form_result.setCurrentIndex(0)
        self.form_profit.clear()

    def save_bet(self):
        """Validate and save / update a bet."""
        sport = self.form_sport.currentText().strip()
        tournament = self.form_tournament.currentText().strip()
        provider = self.form_provider.currentText().strip()
        team_a = self.form_matchup_team_a.currentText().strip()
        team_b = self.form_matchup_team_b.currentText().strip()
        matchup = f"{team_a} vs {team_b}" if team_a and team_b else (team_a or team_b)
        bet = self.form_bet.text().strip()
        live_status = self.form_live_status.currentText()
        odds_text = self.form_odds.text().strip()
        amount_text = self.form_bet_amount.text().strip()
        result_text = self.form_result.currentText()
        profit_text = self.form_profit.text().strip()

        # --- Validation ---
        errors: List[str] = []
        if not sport:
            errors.append("Sport is required.")
        if not tournament:
            errors.append("Tournament is required.")
        if not provider:
            errors.append("Provider is required.")
        if not matchup:
            errors.append("Matchup is required.")
        if not bet:
            errors.append("Bet is required.")

        # Odds
        odds: Optional[float] = None
        try:
            odds = float(odds_text.replace(",", "."))
            if odds <= 0:
                errors.append("Odds must be > 0.")
        except ValueError:
            errors.append("Odds must be a valid positive number.")

        # Bet amount (optional but must be valid if given)
        bet_amount: Optional[float] = None
        if amount_text:
            try:
                bet_amount = float(amount_text.replace(",", "."))
                if bet_amount < 0:
                    errors.append("Bet Amount must be >= 0.")
            except ValueError:
                errors.append("Bet Amount must be a valid number.")

        # Result / Profit (profit is auto-calculated)
        is_pending = (result_text == "(Pending)")
        result: Optional[str] = None
        profit: Optional[float] = None

        if not is_pending:
            result = result_text
            if not profit_text:
                if not bet_amount:
                    errors.append("Bet Amount is required to calculate profit.")
                else:
                    errors.append("Could not calculate profit – check Odds and Bet Amount.")
            else:
                try:
                    profit = float(profit_text.replace(",", "."))
                except ValueError:
                    errors.append("Profit must be a valid number.")

        # If we're editing an existing bet in settle-mode, result must be set.
        # For edit-mode (keeping Pending), we allow saving changes while result remains Pending.

        if errors:
            QMessageBox.warning(self, "Validation Error", "\n".join(errors))
            return

        # --- Save ---
        try:
            if self.editing_bet_id is not None:
                editing_source_tab = getattr(self, "editing_bet_source_tab", 0)
                # If result is still Pending -> update editable fields only
                if is_pending:
                    # Build a fields dict to patch
                    fields: Dict[str, object] = {
                        "sport": sport,
                        "tournament": tournament,
                        "provider": provider,
                        "matchup": matchup,
                        "bet": bet,
                        "live_status": live_status,
                        "odds": float(odds) if odds is not None else None,
                        "bet_amount": float(bet_amount) if bet_amount is not None else None,
                    }
                    if editing_source_tab == 1:
                        fields["result"] = ""
                        fields["profit"] = 0.0
                    # Remove None values
                    fields = {k: v for k, v in fields.items() if v is not None}
                    self.db.update_bet_fields(self.editing_bet_id, fields)
                    QMessageBox.information(self, "Success", "Pending bet updated successfully!")
                    self._reset_form_to_add_mode()
                    self.refresh_data(force=True)
                    self.show_pending_bets_panel()
                    if editing_source_tab == 1 and hasattr(self, "pending_tabs"):
                        self.pending_tabs.setCurrentIndex(1)
                        self.refresh_settled_bets_table()
                else:
                    # Settling the bet (set result & profit)
                    self.db.update_bet(self.editing_bet_id, result, profit)  # type: ignore[arg-type]
                    QMessageBox.information(self, "Success", "Bet settled successfully!")
                    self._reset_form_to_add_mode()
                    self.refresh_data(force=True)
                    self.show_pending_bets_panel()
                    if editing_source_tab == 1 and hasattr(self, "pending_tabs"):
                        self.pending_tabs.setCurrentIndex(1)
                        self.refresh_settled_bets_table()
            else:
                self.db.insert_bet(
                    sport=sport,
                    tournament=tournament,
                    matchup=matchup,
                    bet=bet,
                    live_status=live_status,
                    provider=provider,
                    odds=odds,  # type: ignore[arg-type]
                    bet_amount=bet_amount,
                    result=result,
                    profit=profit,
                )
                msg = "Bet added successfully!" if result else "Pending bet saved!"
                QMessageBox.information(self, "Success", msg)
                self.clear_bet_form()
                self.refresh_data(force=True)
                self.show_data_table_panel()
        except Exception as e:
            QMessageBox.critical(self, "Database Error", f"Failed to save bet:\n{e}")

    # ------------------------------------------------------------------
    # Pending Bets Panel
    # ------------------------------------------------------------------
    def _format_bet_date(self, date_value: str) -> str:
        if not date_value:
            return "N/A"
        try:
            dt = datetime.fromisoformat(str(date_value).replace("Z", "+00:00"))
            return dt.strftime("%d.%m.%Y, %H:%M")
        except ValueError:
            return str(date_value)

    def _create_bet_card(self, row, show_actions: bool = True):
        card = QFrame()
        card.setObjectName("MainCard")
        is_dark = self.dark_mode
        bg_col = "#313244" if is_dark else "#e5e9f0"
        inner_bg_col = "#45475a" if is_dark else "#bcc0cc"
        border_col = "#45475a" if is_dark else "#bcc0cc"
        text_col = "#cdd6f4" if is_dark else "#4c4f69"

        card.setStyleSheet(f"""
            QFrame#MainCard {{
                background-color: {bg_col};
                border: 1px solid {border_col};
                border-radius: 8px;
                padding: 8px;
            }}
            QLabel {{ border: none; color: {text_col}; background: transparent; }}
            QFrame#InnerCard {{
                background-color: {inner_bg_col};
                border-radius: 6px;
                padding: 4px;
                border: none;
            }}
        """)

        c_layout = QVBoxLayout(card)
        c_layout.setContentsMargins(10, 10, 10, 10)
        c_layout.setSpacing(6)

        def make_wrap_label(text):
            lbl = QLabel(text)
            lbl.setWordWrap(True)
            return lbl

        date_str = self._format_bet_date(row[10] if len(row) > 10 else "")
        result_text = str(row[8]) if len(row) > 8 and row[8] else "Pending"

        h_row = QHBoxLayout()
        lbl_date = QLabel(f"<b>Date:</b> {date_str}")
        status_label = "Result" if not show_actions else "Status"
        lbl_status = QLabel(f"<b>{status_label}:</b> {result_text}")
        h_row.addWidget(lbl_date)
        h_row.addStretch(1)
        h_row.addWidget(lbl_status)
        c_layout.addLayout(h_row)

        inner_card = QFrame()
        inner_card.setObjectName("InnerCard")
        inner_layout = QVBoxLayout(inner_card)
        inner_layout.setContentsMargins(6, 6, 6, 6)
        inner_layout.setSpacing(4)

        lbl_sport = make_wrap_label(f"<b>Sport:</b> {row[1] if len(row) > 1 else ''}")
        lbl_tourney = make_wrap_label(f"<b>Tournament:</b> {row[2] if len(row) > 2 else ''}")
        lbl_provider = make_wrap_label(f"<b>Provider:</b> {row[11] if len(row) > 11 and row[11] else 'Unspecified'}")
        lbl_matchup = make_wrap_label(f"{row[3] if len(row) > 3 else ''}")

        inner_layout.addWidget(lbl_sport)
        inner_layout.addWidget(lbl_tourney)
        inner_layout.addWidget(lbl_provider)
        inner_layout.addWidget(lbl_matchup)

        c_layout.addWidget(inner_card)

        lbl_bet = make_wrap_label(f"<b>{row[4] if len(row) > 4 else ''}</b>")
        lbl_bet.setStyleSheet(f"font-size: 15px; color: {text_col}; background: transparent;")
        c_layout.addWidget(lbl_bet)

        bet_row = QHBoxLayout()
        odds_value = float(row[6] or 0.0) if len(row) > 6 else 0.0
        amount_value = float(row[7] or 0.0) if len(row) > 7 else 0.0
        profit_value = float(row[9] or 0.0) if len(row) > 9 else 0.0
        lbl_odds = make_wrap_label(f"<b>Odds:</b> {odds_value:.2f}" if odds_value else "<b>Odds:</b> N/A")
        if show_actions:
            lbl_amount = make_wrap_label(f"<b>Amount:</b> {amount_value:.2f}" if amount_value else "<b>Amount:</b> N/A")
            bet_row.addWidget(lbl_odds)
            bet_row.addWidget(lbl_amount)
        else:
            lbl_amount = make_wrap_label(f"<b>Amount:</b> {amount_value:.2f}" if amount_value else "<b>Amount:</b> N/A")
            lbl_profit = make_wrap_label(f"<b>Profit:</b> {profit_value:.2f}" if profit_value or profit_value == 0 else "<b>Profit:</b> N/A")
            bet_row.addWidget(lbl_odds)
            bet_row.addWidget(lbl_amount)
            bet_row.addWidget(lbl_profit)
        bet_row.addStretch(1)
        c_layout.addLayout(bet_row)

        if show_actions:
            btn_row = QHBoxLayout()
            btn_row.addStretch(1)

            btn_edit = QPushButton("Edit")
            btn_edit.setMinimumWidth(60)
            btn_edit.setStyleSheet(
                "QPushButton { background-color: #ffd166; color: #1e1e2e; font-weight: bold; padding: 4px 16px; border-radius: 6px; border: none; }"
                "QPushButton:hover { background-color: #ffcf80; }"
            )
            btn_edit.clicked.connect(lambda checked, b=row: self._edit_pending_bet(b))

            btn_settle = QPushButton("Settle")
            btn_settle.setMinimumWidth(70)
            btn_settle.setStyleSheet(
                "QPushButton { background-color: #89b4fa; color: #1e1e2e; font-weight: bold; padding: 4px 16px; border-radius: 6px; border: none; }"
                "QPushButton:hover { background-color: #74a8f7; }"
            )
            btn_settle.clicked.connect(lambda checked, b=row: self._settle_bet(b))

            btn_row.addWidget(btn_edit)
            btn_row.addWidget(btn_settle)
            c_layout.addLayout(btn_row)

        return card

    def _populate_bet_cards_layout(self, layout, rows, show_actions: bool = True):
        while layout.count():
            item = layout.takeAt(0)
            if item.widget():
                item.widget().deleteLater()

        layout.setRowStretch(len(rows) // 2 + 1, 1)

        for idx, row in enumerate(rows):
            card = self._create_bet_card(row, show_actions=show_actions)
            row_idx = idx // 2
            col_idx = idx % 2
            layout.addWidget(card, row_idx, col_idx)

    def _load_settled_history_rows(self):
        try:
            rows = [row for row in self.db.fetch_all_bets() if len(row) > 8 and row[8]]
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to load settled bets:\n{e}")
            return []

        rows.sort(key=lambda row: row[10] if len(row) > 10 else "", reverse=True)
        return rows

    def _update_settled_history_view(self):
        rows = getattr(self, "settled_history_rows", [])
        if not hasattr(self, "settled_history_table"):
            return

        total_rows = len(rows)
        rows_per_page = 10
        total_pages = max(1, (total_rows + rows_per_page - 1) // rows_per_page)
        self.settled_history_page_index = max(0, min(getattr(self, "settled_history_page_index", 0), total_pages - 1))

        start = self.settled_history_page_index * rows_per_page
        end = min(start + rows_per_page, total_rows)
        page_rows = rows[start:end]

        self.lbl_settled_count.setText(f"{total_rows} settled bet(s)")
        self.lbl_settled_page.setText(f"Showing {start + 1 if total_rows else 0}-{end} of {total_rows}")
        self.btn_settled_prev.setEnabled(self.settled_history_page_index > 0)
        self.btn_settled_next.setEnabled(self.settled_history_page_index < total_pages - 1)

        self.settled_history_table.blockSignals(True)
        self.settled_history_table.setRowCount(len(page_rows))

        headers = ["Date", "Sport", "Tournament", "Matchup", "Bet", "Live", "Provider", "Odds", "Result", "Profit"]
        for row_idx, row in enumerate(page_rows):
            values = [
                self._format_bet_date(row[10] if len(row) > 10 else ""),
                str(row[1] if len(row) > 1 else ""),
                str(row[2] if len(row) > 2 else ""),
                str(row[3] if len(row) > 3 else ""),
                str(row[4] if len(row) > 4 else ""),
                str(row[5] if len(row) > 5 else ""),
                str(row[11] if len(row) > 11 and row[11] else "Unspecified"),
                f"{float(row[6] or 0.0):.2f}" if len(row) > 6 else "N/A",
                str(row[8] if len(row) > 8 else ""),
                f"{float(row[9] or 0.0):.2f}" if len(row) > 9 else "0.00",
            ]
            for col_idx, value in enumerate(values):
                item = QTableWidgetItem(value)
                if col_idx in (7, 9):
                    item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
                self.settled_history_table.setItem(row_idx, col_idx, item)

        self.settled_history_table.blockSignals(False)
        self.settled_history_table.resizeRowsToContents()
        self._sync_history_edit_button_state()

    def _build_pending_bets_panel(self, parent_layout):
        """Build the pending-bets view with a scroll area of cards and settle buttons."""
        self.pending_bets_panel = QWidget()
        p_layout = QVBoxLayout(self.pending_bets_panel)
        p_layout.setContentsMargins(10, 10, 10, 10)

        self.pending_tabs = QTabWidget()
        self.pending_tabs.currentChanged.connect(self._refresh_pending_bets_tab)

        pending_page = QWidget()
        pending_page_layout = QVBoxLayout(pending_page)
        pending_page_layout.setContentsMargins(0, 0, 0, 0)

        self.lbl_pending_count = QLabel("0 pending bet(s)")
        self.lbl_pending_count.setStyleSheet("font-size: 15px; font-weight: bold;")
        pending_page_layout.addWidget(self.lbl_pending_count)

        self.pending_scroll = QScrollArea()
        self.pending_scroll.setWidgetResizable(True)
        self.pending_scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        self.pending_scroll.setStyleSheet("QScrollArea { border: none; }")

        self.pending_cards_container = QWidget()
        self.pending_cards_layout = QGridLayout(self.pending_cards_container)
        self.pending_cards_layout.setSpacing(10)
        self.pending_cards_layout.setContentsMargins(0, 0, 0, 0)
        self.pending_cards_layout.setColumnStretch(0, 1)
        self.pending_cards_layout.setColumnStretch(1, 1)

        self.pending_scroll.setWidget(self.pending_cards_container)
        pending_page_layout.addWidget(self.pending_scroll, 1)

        self.settled_page = QWidget()
        settled_page_layout = QVBoxLayout(self.settled_page)
        settled_page_layout.setContentsMargins(0, 0, 0, 0)
        settled_page_layout.setSpacing(8)

        settled_header = QHBoxLayout()
        self.lbl_settled_count = QLabel("0 settled bet(s)")
        self.lbl_settled_count.setStyleSheet("font-size: 15px; font-weight: bold;")
        self.lbl_settled_page = QLabel("Showing 0-0 of 0")
        self.lbl_settled_page.setStyleSheet("font-size: 13px; color: #8b90a0;")
        settled_header.addWidget(self.lbl_settled_count)
        settled_header.addStretch(1)
        settled_header.addWidget(self.lbl_settled_page)
        settled_page_layout.addLayout(settled_header)

        settled_nav = QHBoxLayout()
        self.btn_settled_prev = QPushButton("Previous 10")
        self.btn_settled_prev.clicked.connect(lambda checked=False: self._change_settled_history_page(-1))
        self.btn_settled_next = QPushButton("Next 10")
        self.btn_settled_next.clicked.connect(lambda checked=False: self._change_settled_history_page(1))
        settled_nav.addWidget(self.btn_settled_prev)
        settled_nav.addWidget(self.btn_settled_next)
        self.btn_settled_edit = QPushButton("Edit Selected")
        self.btn_settled_edit.setEnabled(False)
        self.btn_settled_edit.clicked.connect(self._edit_selected_history_bet)
        settled_nav.addWidget(self.btn_settled_edit)
        settled_nav.addStretch(1)
        settled_page_layout.addLayout(settled_nav)

        self.settled_history_table = QTableWidget(0, 10)
        self.settled_history_table.setHorizontalHeaderLabels(["Date", "Sport", "Tournament", "Matchup", "Bet", "Live", "Provider", "Odds", "Result", "Profit"])
        self.settled_history_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.settled_history_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.settled_history_table.setAlternatingRowColors(True)
        self.settled_history_table.verticalHeader().setVisible(False)
        self.settled_history_table.setShowGrid(False)
        self.settled_history_table.setWordWrap(True)
        self.settled_history_table.cellDoubleClicked.connect(self._edit_selected_history_bet)
        self.settled_history_table.itemSelectionChanged.connect(self._sync_history_edit_button_state)
        history_header = self.settled_history_table.horizontalHeader()
        history_header.setStretchLastSection(False)
        history_header.setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
        history_header.setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)
        history_header.setSectionResizeMode(2, QHeaderView.ResizeMode.Stretch)
        history_header.setSectionResizeMode(3, QHeaderView.ResizeMode.Stretch)
        history_header.setSectionResizeMode(4, QHeaderView.ResizeMode.Stretch)
        history_header.setSectionResizeMode(5, QHeaderView.ResizeMode.ResizeToContents)
        history_header.setSectionResizeMode(6, QHeaderView.ResizeMode.ResizeToContents)
        history_header.setSectionResizeMode(7, QHeaderView.ResizeMode.ResizeToContents)
        history_header.setSectionResizeMode(8, QHeaderView.ResizeMode.ResizeToContents)
        history_header.setSectionResizeMode(9, QHeaderView.ResizeMode.ResizeToContents)
        settled_page_layout.addWidget(self.settled_history_table, 1)

        self.pending_tabs.addTab(pending_page, "Pending Bets")
        self.pending_tabs.addTab(self.settled_page, "History")
        p_layout.addWidget(self.pending_tabs, 1)

        parent_layout.addWidget(self.pending_bets_panel, 1)
        self.pending_bets_panel.hide()

    def _refresh_pending_bets_tab(self, index: int):
        if index == 1:
            self.refresh_settled_bets_table()
        else:
            self.refresh_pending_bets_table()

    def _change_settled_history_page(self, delta: int):
        self.settled_history_page_index = max(0, getattr(self, "settled_history_page_index", 0) + delta)
        self._update_settled_history_view()

    def _sync_history_edit_button_state(self):
        if hasattr(self, "btn_settled_edit") and hasattr(self, "settled_history_table"):
            self.btn_settled_edit.setEnabled(self.settled_history_table.currentRow() >= 0)

    def _edit_selected_history_bet(self, *args):
        if not hasattr(self, "settled_history_table"):
            return

        table_row = self.settled_history_table.currentRow()
        if table_row < 0:
            return

        rows = getattr(self, "settled_history_rows", [])
        page_index = getattr(self, "settled_history_page_index", 0)
        source_index = page_index * 10 + table_row
        if source_index < 0 or source_index >= len(rows):
            return

        self._edit_bet(rows[source_index], source_tab=1, preserve_result=True)

    def show_pending_bets_panel(self):
        """Switch to the pending-bets view."""
        self.current_view = "pending_bets"
        self._update_nav_buttons(self.btn_pending_bets)
        self.table.hide()
        self.statistics_panel.hide()
        self.add_bet_panel.hide()
        self.pending_bets_panel.show()
        if hasattr(self, "pending_tabs"):
            self.pending_tabs.setCurrentIndex(0)
        self.settled_history_page_index = 0
        self.refresh_pending_bets_table()
        self.refresh_settled_bets_table()

    def refresh_pending_bets_table(self):
        """Fetch and display pending bets as cards."""
        try:
            rows = self.db.fetch_pending_bets()
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to load pending bets:\n{e}")
            return
        
        self.lbl_pending_count.setText(f"{len(rows)} pending bet(s)")
        self._populate_bet_cards_layout(self.pending_cards_layout, rows, show_actions=True)

    def refresh_settled_bets_table(self):
        """Fetch and display settled bets as a paginated line list."""
        self.settled_history_rows = self._load_settled_history_rows()
        if not hasattr(self, "settled_history_page_index"):
            self.settled_history_page_index = 0
        self._update_settled_history_view()

    def _settle_bet(self, row_data):
        """Open a small dialog to quickly settle a pending bet (Win / Lose)."""
        # row_data: (id, sport, tournament, matchup, bet,
        #            live_status, odds, bet_amount, result, profit, date_created, provider)
        bid = row_data[0]
        sport = str(row_data[1])
        tournament = str(row_data[2])
        matchup = str(row_data[3])
        bet_text = str(row_data[4])
        odds = float(row_data[6] or 0.0)
        bet_amount = float(row_data[7] or 0.0)

        dlg = QDialog(self)
        dlg.setWindowTitle("Settle Bet")
        dlg.setModal(True)
        dlg.setFixedSize(520, 220)
        layout = QVBoxLayout(dlg)

        amount_str = f"{bet_amount:.2f}" if bet_amount else "N/A"
        info = QLabel(
            f"<b>Sport:</b> {sport}<br>"
            f"<b>Tournament:</b> {tournament}<br>"
            f"<b>Provider:</b> {str(row_data[11] if len(row_data) > 11 and row_data[11] else 'Unspecified')}<br>"
            f"<b>Matchup:</b> {matchup}<br>"
            f"<b>Bet:</b> {bet_text}<br>"
            f"<b>Odds:</b> {odds:.2f}   <b>Amount:</b> {amount_str}"
        )
        info.setWordWrap(True)
        layout.addWidget(info)

        btn_row = QHBoxLayout()
        btn_row.addStretch(1)

        btn_win = QPushButton("Win")
        btn_win.setStyleSheet("QPushButton { background-color: #2ecc71; color: white; font-weight: bold; padding: 6px 18px; border-radius: 6px; }")
        btn_lose = QPushButton("Lose")
        btn_lose.setStyleSheet("QPushButton { background-color: #e74c3c; color: white; font-weight: bold; padding: 6px 18px; border-radius: 6px; }")
        btn_void = QPushButton("Delete Bet")
        btn_void.setStyleSheet("QPushButton { background-color: #f39c12; color: white; font-weight: bold; padding: 6px 18px; border-radius: 6px; }")
        btn_cancel = QPushButton("Close")
        btn_cancel.setStyleSheet("QPushButton { padding: 6px 12px; }")

        def ensure_amount() -> Optional[float]:
            nonlocal bet_amount
            if bet_amount and bet_amount > 0:
                return bet_amount
            val, ok = QInputDialog.getDouble(self, "Bet Amount", "Enter bet amount:", 0.0, 0.0, 1000000.0, 2)
            if not ok:
                return None
            bet_amount = float(val)
            return bet_amount

        def do_settle(is_win: bool):
            amt = ensure_amount()
            if amt is None:
                return
            if is_win:
                profit = round(odds * amt - amt, 2)
                result = "Win"
            else:
                profit = round(-amt, 2)
                result = "Lose"
            try:
                self.db.update_bet(bid, result, profit)
                QMessageBox.information(self, "Success", f"Bet settled as {result} (Profit: {profit:.2f})")
                dlg.accept()
                # Refresh UI
                self.refresh_data(force=True)
                self.show_pending_bets_panel()
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Failed to settle bet:\n{e}")

        def do_delete():
            # Delete the bet completely
            try:
                self.db.delete_bet(bid)
                QMessageBox.information(self, "Success", "Bet has been removed.")
                dlg.accept()
                self.refresh_data(force=True)
                self.show_pending_bets_panel()
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Failed to remove bet:\n{e}")

        btn_win.clicked.connect(lambda checked: do_settle(True))
        btn_lose.clicked.connect(lambda checked: do_settle(False))
        btn_void.clicked.connect(lambda checked: do_delete())
        btn_cancel.clicked.connect(dlg.reject)

        btn_row.addWidget(btn_win)
        btn_row.addWidget(btn_lose)
        btn_row.addWidget(btn_void)
        btn_row.addWidget(btn_cancel)
        layout.addLayout(btn_row)

        dlg.exec()

    def _edit_bet(self, row_data, source_tab: int = 0, preserve_result: bool = False):
        """Open the add-bet form in editable mode for a bet row."""
        # row_data: (id, sport, tournament, matchup, bet,
        #            live_status, odds, bet_amount, result, profit, date_created, provider)
        self.editing_bet_id = row_data[0]
        self.editing_bet_source_tab = source_tab
        self.current_view = "add_bet"
        self._update_nav_buttons(self.btn_add_bet)
        self.table.hide()
        self.statistics_panel.hide()
        self.pending_bets_panel.hide()
        self.add_bet_panel.show()

        self.add_bet_title.setText("Edit Bet")
        self.btn_save_bet.setText("Update Bet")

        self._populate_form_sports()
        self._populate_form_providers()

        # Fill fields (keep all editable)
        self.form_sport.setCurrentText(str(row_data[1]))
        self.form_tournament.setCurrentText(str(row_data[2]))
        if len(row_data) > 11:
            self.form_provider.setCurrentText(str(row_data[11]))
        
        matchup_str = str(row_data[3])
        team_a, team_b = "", ""
        if " vs " in matchup_str:
            parts = matchup_str.split(" vs ", 1)
            team_a = parts[0].strip()
            if len(parts) > 1:
                team_b = parts[1].strip()
        else:
            team_a = matchup_str
            
        self.form_matchup_team_a.setCurrentText(team_a)
        self.form_matchup_team_b.setCurrentText(team_b)
        
        self.form_bet.setText(str(row_data[4]))
        idx = self.form_live_status.findText(str(row_data[5]))
        if idx >= 0:
            self.form_live_status.setCurrentIndex(idx)
        self.form_odds.setText(f"{row_data[6]:.2f}" if row_data[6] else "")
        self.form_bet_amount.setText(f"{row_data[7]:.2f}" if row_data[7] else "")
        if preserve_result and len(row_data) > 8 and row_data[8]:
            result_text = str(row_data[8])
            result_index = self.form_result.findText(result_text)
            if result_index >= 0:
                self.form_result.setCurrentIndex(result_index)
            else:
                self.form_result.setCurrentIndex(0)
            self.form_profit.setText(f"{row_data[9]:.2f}" if len(row_data) > 9 and row_data[9] is not None else "")
        else:
            self.form_result.setCurrentIndex(0)  # keep as Pending
            self.form_profit.clear()

        # Ensure fields are enabled for editing
        for w in [self.form_sport, self.form_tournament, 
                self.form_provider, self.form_matchup_team_a, self.form_matchup_team_b,
                  self.form_bet, self.form_live_status,
                  self.form_odds, self.form_bet_amount]:
            w.setEnabled(True)

        self.form_result.setEnabled(True)
        self.form_profit.setReadOnly(True)

    def _edit_pending_bet(self, row_data):
        self._edit_bet(row_data, source_tab=0, preserve_result=False)

    def dismiss_changes_panel(self):
        """Hide the changes notification panel"""
        self.changes_panel.hide()

    def _detect_odds_changes(self) -> List[Dict]:
        """Detect changes at the odds level showing previous vs new EV, WR, and count"""
        if not self._previous_data_cache:
            return []  # No previous data to compare against
        
        changes = []
        
        # Check each sport in the new data
        for sport, new_entry in self.data_cache.items():
            old_entry = self._previous_data_cache.get(sport)
            
            # Build dict of old odds data for quick lookup
            old_odds_map = {}
            if old_entry:
                for row in old_entry.rows:
                    odds = row[0]  # odds is first element
                    old_odds_map[odds] = {
                        'live_wr': row[1],
                        'prem_wr': row[2], 
                        'live_cnt': row[3],
                        'prem_cnt': row[4]
                    }
            
            # Check each new odds entry
            for row in new_entry.rows:
                odds, live_wr, prem_wr, live_cnt, prem_cnt = row
                
                if odds not in old_odds_map:
                    # New odds entry - check which type has data
                    has_live = live_cnt is not None and live_cnt > 0
                    has_prem = prem_cnt is not None and prem_cnt > 0
                    
                    if has_live:
                        changes.append({
                            'sport': sport,
                            'odds': odds,
                            'bet_type': 'LIVE',
                            'type': 'new',
                            'prev_wr': None,
                            'prev_cnt': None,
                            'new_wr': live_wr,
                            'new_cnt': live_cnt
                        })
                    if has_prem:
                        changes.append({
                            'sport': sport,
                            'odds': odds,
                            'bet_type': 'PREMATCH',
                            'type': 'new',
                            'prev_wr': None,
                            'prev_cnt': None,
                            'new_wr': prem_wr,
                            'new_cnt': prem_cnt
                        })
                else:
                    # Existing odds - check which type changed
                    old = old_odds_map[odds]
                    
                    # Check LIVE changes
                    if old['live_cnt'] != live_cnt or old['live_wr'] != live_wr:
                        changes.append({
                            'sport': sport,
                            'odds': odds,
                            'bet_type': 'LIVE',
                            'type': 'updated',
                            'prev_wr': old['live_wr'],
                            'prev_cnt': old['live_cnt'],
                            'new_wr': live_wr,
                            'new_cnt': live_cnt
                        })
                    
                    # Check PREMATCH changes
                    if old['prem_cnt'] != prem_cnt or old['prem_wr'] != prem_wr:
                        changes.append({
                            'sport': sport,
                            'odds': odds,
                            'bet_type': 'PREMATCH',
                            'type': 'updated',
                            'prev_wr': old['prem_wr'],
                            'prev_cnt': old['prem_cnt'],
                            'new_wr': prem_wr,
                            'new_cnt': prem_cnt
                        })
        
        return changes

    def _calc_ev_display(self, wr: Optional[float], odds: float, cnt: Optional[int]) -> Tuple[str, Optional[float]]:
        """Calculate EV for display"""
        if wr is None or cnt is None:
            return "N/A", None
        ev = calc_ev(odds, wr, cnt)
        if ev is None:
            return "N/A", None
        return f"{ev*100:.2f}%", ev

    def _format_odds_change(self, change: Dict) -> str:
        """Format a single odds change for display - showing only the specific bet type that changed"""
        sport = change['sport']
        odds = change['odds']
        ctype = change['type']
        bet_type = change['bet_type']
        
        icon = "🔴" if bet_type == "LIVE" else "⚫"
        
        prev_ev_str, prev_ev = self._calc_ev_display(change['prev_wr'], odds, change['prev_cnt'])
        new_ev_str, new_ev = self._calc_ev_display(change['new_wr'], odds, change['new_cnt'])
        
        lines = [f"<b>{sport}</b> | Odds: {odds:.2f} | {icon} {bet_type} ({ctype.upper()})<br>"]
        
        if change['prev_cnt'] is not None:
            lines.append(f"&nbsp;&nbsp;Previous: WR={fmt_wr(change['prev_wr'])}, Count={change['prev_cnt']}, EV={prev_ev_str}<br>")
        else:
            lines.append(f"&nbsp;&nbsp;Previous: No data<br>")
        
        lines.append(f"&nbsp;&nbsp;<b>Updated:</b> WR={fmt_wr(change['new_wr'])}, Count={change['new_cnt']}, EV={new_ev_str}<br>")
        
        if change['prev_cnt'] is not None and prev_ev is not None and new_ev is not None:
            diff = new_ev - prev_ev
            diff_str = f"+{diff*100:.2f}%" if diff >= 0 else f"{diff*100:.2f}%"
            color = "green" if diff >= 0 else "red"
            lines.append(f"&nbsp;&nbsp;EV Change: <span style='color:{color}'>{diff_str}</span><br>")
        
        lines.append("<br>")
        return "".join(lines)

    def _format_changes_message(self, changes: List[Dict]) -> str:
        """Format the list of odds changes into a readable message"""
        if not changes:
            return "No new data was added."
        
        lines = [f"<b>{len(changes)} odds value(s) updated:</b><br><br>"]
        
        for change in changes:
            lines.append(self._format_odds_change(change))
        
        return "".join(lines)

    def _display_changes(self, changes: List[Dict]):
        """Display the changes panel with detected changes only"""
        if not changes:
            self.changes_panel.hide()
            return
        
        message = self._format_changes_message(changes)
        count = len(changes)
        self.changes_title.setText(f"📊 {count} Odds Value(s) Changed")
        self.changes_text.setText(message)
        self.changes_panel.show()

    def _format_current_data_message(self) -> str:
        """Format message showing current data state"""
        total_odds = sum(len(entry.rows) for entry in self.data_cache.values())
        total_bets = len(self.matchbet_data)
        
        lines = [f"<b>Data Loaded Successfully</b><br><br>"]
        lines.append(f"Total Sports: {len(self.data_cache)}<br>")
        lines.append(f"Total Odds Values: {total_odds}<br>")
        lines.append(f"Total Bets: {total_bets}<br><br>")
        
        # Show summary by sport
        for sport in sorted(self.data_cache.keys()):
            entry = self.data_cache[sport]
            sport_bets = sum((r[3] or 0) + (r[4] or 0) for r in entry.rows)
            lines.append(f"• <b>{sport}</b>: {len(entry.rows)} odds values ({sport_bets} bets)<br>")
        
        return "".join(lines)

    def refresh_data(self, force: bool=False, force_network: bool=False):
        sheet = self.sport_combo.currentText()
        # If we already have this sheet cached and not forcing, use cache instantly
        if not force and sheet in self.data_cache:
            self.setWindowTitle(f"EV Bet Calculator - {sheet}")
            self.set_status("Loaded from cache")
            self._apply_filters()
            return
        # Otherwise perform an async refresh from database
        # Store previous data for comparison (always track changes)
        self._previous_data_cache = self.data_cache.copy()
        self._previous_matchbet_data = self.matchbet_data.copy()
        self.set_status(f"Refreshing all data..."); self.set_controls_enabled(False)
        t=QThread(); w=RefreshWorker(self.db, force_refresh_network=force_network); w.moveToThread(t)
        t.started.connect(w.run)
        w.finished.connect(lambda ok, info, res: self._on_refresh_done(t,w,ok,info,res))
        t.start()

    def _on_refresh_done(self, thread: QThread, worker: RefreshWorker, ok: bool, info: str, res: object):
        thread.quit(); thread.wait(); worker.deleteLater()
        if ok and res:
            cache, data = res
            self.data_cache = cache
            
            # Detect and display changes at odds level
            changes = self._detect_odds_changes()
            self._display_changes(changes)
            
            self.matchbet_data = data
            
            # Update sport combo if new sports appeared
            current_sport = self.sport_combo.currentText()
            self.sport_combo.blockSignals(True)
            self.sport_combo.clear()
            self.sport_combo.addItems(self.get_sorted_sports())
            if current_sport in self.data_cache:
                self.sport_combo.setCurrentText(current_sport)
            elif self.sport_combo.count() > 0:
                self.sport_combo.setCurrentIndex(0)
            self.sport_combo.blockSignals(False)
            
            # Repopulate tournament and team combos for the current sport
            current_sport = self.sport_combo.currentText()
            self._populate_tournament_combo(current_sport)
            self._populate_team_combo(current_sport)
            
            # Refresh view using filters
            if current_sport in self.data_cache:
                self.setWindowTitle(f"EV Bet Calculator - {current_sport}")
                self._apply_filters()
            else:
                # Cache is empty (e.g. after DB reset) – clear visible tables
                self.setWindowTitle("EV Bet Calculator")
                self.table.setRowCount(0)
                self.refresh_pending_bets_table()
        else:
            QMessageBox.critical(self, "Error", f"Failed to load data: {info}")
        self.set_controls_enabled(True); self.set_status("Ready")

    def set_initial_cache(self, cache: Dict[str, SheetCacheEntry]):
        self.data_cache = cache
        
        # Populate sports combo
        self.sport_combo.blockSignals(True)
        self.sport_combo.clear()
        sports = self.get_sorted_sports()
        self.sport_combo.addItems(sports)
        self.sport_combo.blockSignals(False)
        
        if sports:
            first = sports[0]
            self.sport_combo.setCurrentText(first)
            self._populate_tournament_combo(first)
            self._populate_team_combo(first)
            self._populate_provider_combo(first)
            self._apply_filters()

    def set_matchbet_data(self, data: List[MatchBetTuple]):
        self.matchbet_data = data
        if hasattr(self, "sport_combo"):
            self._populate_provider_combo(self.sport_combo.currentText())


def main():
    app = QApplication(sys.argv)

    # Ensure the database is created
    db = DatabaseManager()
    db.create_tables()

    win = MainWindow()
    dlg = PreloadDialog()
    preload_thread = QThread()
    worker = PreloadWorker(win.db)
    worker.moveToThread(preload_thread)
    preload_thread.started.connect(worker.run)
    worker.progress.connect(dlg.update_progress)
    worker.status.connect(dlg.update_status)

    def done(ok: bool, msg: str):
        preload_thread.quit(); preload_thread.wait(); worker.deleteLater(); dlg.accept()
        if not ok:
            QMessageBox.critical(win, "Preload Failed", f"Failed to preload data:\n{msg}")
            win.close()
            return
        win.set_initial_cache(worker.cache)
        win.set_matchbet_data(worker.matchbet_data)
        win.show()
        QTimer.singleShot(0, lambda: win.refresh_data(force=True))

    worker.finished.connect(done)
    QTimer.singleShot(60000, lambda: done(False, "Preloading timed out") if preload_thread.isRunning() else None)
    preload_thread.start()
    dlg.exec()
    return app.exec()


if __name__ == '__main__':
    sys.exit(main())