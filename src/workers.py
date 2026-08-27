from __future__ import annotations

from typing import List, Dict

from PyQt6.QtCore import QObject, pyqtSignal

try:
    from src.database import DatabaseManager
except ModuleNotFoundError:
    from database import DatabaseManager

try:
    from src.models import MatchBetTuple, SheetCacheEntry
except ModuleNotFoundError:
    from models import MatchBetTuple, SheetCacheEntry

try:
    from src.data_processing import fetch_matchbet_data_from_db, process_bets_to_cache
except ModuleNotFoundError:
    from data_processing import fetch_matchbet_data_from_db, process_bets_to_cache

class PreloadWorker(QObject):
    progress = pyqtSignal(str)
    status = pyqtSignal(str)
    finished = pyqtSignal(bool, str)
    def __init__(self, db: DatabaseManager):
        super().__init__()
        self.db = db
        self.cache: Dict[str, SheetCacheEntry] = {}
        self.matchbet_data: List[MatchBetTuple] = []

    def run(self):
        try:
            self.progress.emit("Loading data from database...")
            self.status.emit("Querying settled bets...")
            
            self.matchbet_data = fetch_matchbet_data_from_db(self.db)
            self.status.emit(f"Loaded {len(self.matchbet_data)} bets. Processing...")
            
            self.cache = process_bets_to_cache(self.matchbet_data)
            
            self.progress.emit("Finalizing...")
            self.finished.emit(True, "")
        except Exception as e:
            self.finished.emit(False, str(e))


class RefreshWorker(QObject):
    finished = pyqtSignal(bool, str, object)
    def __init__(self, db: DatabaseManager, force_refresh_network: bool = False):
        super().__init__()
        self.db = db
        self.force_refresh_network = force_refresh_network

    def run(self):
        try:
            data = fetch_matchbet_data_from_db(self.db, self.force_refresh_network)
            cache = process_bets_to_cache(data)
            self.finished.emit(True, "Refreshed all data", (cache, data))
        except Exception as e:
            self.finished.emit(False, str(e), None)


class MigrationWorker(QObject):
    """Background worker to import an .xlsx file into the SQLite database."""
    progress = pyqtSignal(int, int)  # current, total
    finished = pyqtSignal(bool, str, int, int, int)  # ok, msg, total, settled, pending

    def __init__(self, file_path: str, db: DatabaseManager):
        super().__init__()
        self.file_path = file_path
        self.db = db

    def run(self):
        try:
            import openpyxl
        except ImportError:
            self.finished.emit(False, "openpyxl is not installed. Run: pip install openpyxl", 0, 0, 0)
            return
        try:
            wb = openpyxl.load_workbook(self.file_path, read_only=True, data_only=True)
            if "MATCHBET" not in wb.sheetnames:
                self.finished.emit(False, "Sheet 'MATCHBET' not found in the workbook.", 0, 0, 0)
                wb.close()
                return
            ws = wb["MATCHBET"]
            rows = list(ws.iter_rows(min_row=2, values_only=True))
            total = len(rows)
            db = self.db
            db.create_tables()
            imported = 0
            settled = 0
            pending = 0
            for i, row in enumerate(rows):
                self.progress.emit(i + 1, total)
                if len(row) < 6:
                    continue
                sport = str(row[0] or "").strip()
                if not sport:
                    continue
                tournament = str(row[1] or "").strip()
                matchup = str(row[2] or "").strip()
                bet = str(row[3] or "").strip()
                live_status = str(row[4] or "").strip() if len(row) > 4 else "NOT LIVE"
                odds_raw = row[5] if len(row) > 5 else None
                bet_amount_raw = row[6] if len(row) > 6 else None
                result_raw = str(row[7] or "").strip() if len(row) > 7 else ""
                profit_raw = row[8] if len(row) > 8 else None
                provider_raw = str(row[9] or "").strip() if len(row) > 9 else ""
                provider = provider_raw or "BetBy"

                try:
                    odds = float(str(odds_raw).replace(",", ".")) if odds_raw is not None else None
                except (ValueError, TypeError):
                    odds = None
                if odds is None or odds <= 0:
                    continue

                try:
                    bet_amount = float(str(bet_amount_raw).replace(",", ".")) if bet_amount_raw not in (None, "") else None
                except (ValueError, TypeError):
                    bet_amount = None

                result = result_raw if result_raw else None
                profit = None
                if result:
                    try:
                        profit = float(str(profit_raw).replace(",", ".")) if profit_raw not in (None, "") else None
                    except (ValueError, TypeError):
                        profit = None
                    settled += 1
                else:
                    pending += 1

                db.insert_bet(
                    sport=sport,
                    tournament=tournament,
                    matchup=matchup,
                    bet=bet,
                    live_status=live_status if live_status else "NOT LIVE",
                    provider=provider,
                    odds=odds,
                    bet_amount=bet_amount,
                    result=result,
                    profit=profit,
                )
                imported += 1
            wb.close()
            self.finished.emit(True, "", imported, settled, pending)
        except Exception as e:
            self.finished.emit(False, str(e), 0, 0, 0)


